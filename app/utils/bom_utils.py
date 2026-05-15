"""
BOM (Bill of Materials) Utility Fonksiyonları
==============================================
ÇELMAK Ürün Ağacı Excel formatı — iki format desteklenir:

FORMAT A — Numaralandırma Kolonu (ekran görüntüsündeki gerçek format):
  • İlk non-null sütunda "1.1.", "1.1.2." gibi noktalı değerler
  • Bölüm başlıkları: "1. KAFA", "2. GÖVDE" (birleştirilmiş/stillendirilmiş satır)
  • Her bölüm altında bir başlık satırı (Numara | Adlandırması | ... atlanır)
  • Parça adı, kod, malzeme cinsi, adet kolonları

FORMAT B — Kolon Girintisi  (urun_agaci_sablon.xlsx şablonu):
  • Col A: Ana ürün adı (sadece 1. satır)
  • Col B: L1 grup (Kasa, Şase, Dingil…)
  • Col C: L2 alt grup
  • Col D: Parça adı
  • Col H: Adet, Col I: Birim
"""

import re
import openpyxl
from decimal import Decimal
import unicodedata
from collections import Counter


# ---------------------------------------------------------------------------
# Ürün kodu üreteci
# ---------------------------------------------------------------------------

_TR_MAP = {'\u00e7': 'C', '\u00c7': 'C', '\u011f': 'G', '\u011e': 'G', '\u0131': 'I', '\u0130': 'I',
           '\u00f6': 'O', '\u00d6': 'O', '\u015f': 'S', '\u015e': 'S', '\u00fc': 'U', '\u00dc': 'U'}

STANDARD_PREFIXES = {'201', '202', '203', '204', '205', '206', '207', '208', '209',
                     '210', '211', '212', '213', '214', '216', '217', '219'}

MATERIAL_WORD_ALIASES = {
    'ST3237': 'ST37',
    'ST37': 'ST37',
    'ST44': 'ST44',
    '4140': 'C4140',
    'S235': 'ST37',
    'SIYAH': 'SIYAH',
    'SAC': 'SAC',
    'LAMA': 'LAMA',
    'BORU': 'BORU',
    'SANAYI': 'SANAYI',
    'PROFIL': 'PROFIL',
    'TRANSMISYON': 'TRANSMISYON',
    'CEKME': 'CEKME',
    'CELIK': 'CELIK',
    'C1040': 'C1040',
    'C4140': 'C4140',
    'Ç1040': 'C1040',
}


def _make_product_code(name: str, bom_code: str = '') -> str:
    """Ürün adından veya mevcut BOM kodundan ürün kodu üretir."""
    if bom_code and bom_code.strip():
        s = bom_code.strip()
    else:
        s = name
    for k, v in _TR_MAP.items():
        s = s.replace(k, v)
    s = re.sub(r'[^A-Za-z0-9]', '-', s).upper()
    s = re.sub(r'-+', '-', s).strip('-')
    return s[:20] or 'BOM'


def _unique_product_code(base: str) -> str:
    """Benzersiz ürün kodu döndürür; gerekirsge sayısal sone ekler."""
    from app.models import Product
    code = base
    suffix = 1
    while Product.query.filter_by(code=code).first():
        code = f'{base[:17]}-{suffix:02d}'
        suffix += 1
    return code

# ---------------------------------------------------------------------------
# Yardımcılar
# ---------------------------------------------------------------------------

def _ascii_upper(value: str) -> str:
    """Turkish-safe uppercase text for fuzzy material matching."""
    text = str(value or '')
    for tr, en in _TR_MAP.items():
        text = text.replace(tr, en)
    text = text.replace('Ø', ' CAP ').replace('ø', ' CAP ')
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()


def _material_tokens(value: str) -> set[str]:
    """Normalize raw material text into comparable tokens."""
    text = _ascii_upper(value)
    replacements = {
        'ST 37': 'ST37',
        'ST-37': 'ST37',
        'ST 44': 'ST44',
        'ST-44': 'ST44',
        'C 1040': 'C1040',
        'C-1040': 'C1040',
        'CAP': ' ',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    raw_tokens = re.findall(r'[A-Z0-9]+(?:X[A-Z0-9]+)*', text)
    ignored = {'MM', 'KG', 'MT', 'M', 'METRE', 'KILOGRAM', 'HAZIR', 'HAMMADDE', 'PARCA', 'PARÇA'}
    tokens = set()
    for token in raw_tokens:
        if token in ignored:
            continue
        tokens.add(MATERIAL_WORD_ALIASES.get(token, token))
    return tokens


def _material_number(value: str) -> str:
    number = str(value or '').replace(',', '.')
    if not number:
        return ''
    try:
        return f'{float(number):g}'
    except ValueError:
        return number.strip()


def _dimension_parts(value: str) -> tuple[str, ...]:
    return tuple(_material_number(part) for part in re.split(r'\s*[Xx]\s*', value) if part)


def _strict_material_signature(value: str):
    """Return a strict comparable material signature for measured raw materials."""
    text = _ascii_upper(value).replace(',', '.')
    tokens = _material_tokens(text)

    family = None
    if 'SAC' in tokens:
        family = 'SAC'
    elif 'LAMA' in tokens:
        family = 'LAMA'
    elif 'PROFIL' in tokens:
        family = 'PROFIL'
    elif tokens & {'BORU', 'SANAYI', 'CELIK'}:
        family = 'BORU'
    elif 'TRANSMISYON' in tokens:
        family = 'TRANSMISYON'
    elif tokens & {'C1040', 'C4140', 'CEKME'}:
        family = 'MIL'

    if not family:
        return None

    grades = tuple(sorted(tokens & {'ST37', 'ST44', 'S235', 'C1040', 'C4140', 'GG25', 'GGG45', 'GGG50'}))
    dimensions = []

    x_matches = re.findall(r'(\d+(?:\.\d+)?\s*[Xx]\s*\d+(?:\.\d+)?(?:\s*[Xx]\s*\d+(?:\.\d+)?)*)', text)
    for match in x_matches:
        dimensions.append('X'.join(_dimension_parts(match)))

    cap_matches = re.findall(r'\bCAP\s*(\d+(?:\.\d+)?)', text)
    for match in cap_matches:
        dimensions.append(f'CAP{_material_number(match)}')

    if family == 'SAC':
        thickness_matches = re.findall(r'(?<![A-Z0-9])(\d+(?:\.\d+)?)\s*MM\b', text)
        if thickness_matches:
            dimensions = [f'T{_material_number(thickness_matches[-1])}']

    if not dimensions:
        return None

    return {
        'family': family,
        'grades': grades,
        'dimensions': tuple(sorted(set(dimensions))),
    }


def _strict_signatures_match(source_sig, candidate_sig) -> bool:
    if not source_sig:
        return True
    if not candidate_sig:
        return False
    if source_sig['family'] != candidate_sig['family']:
        return False
    if source_sig['dimensions'] != candidate_sig['dimensions']:
        return False
    if source_sig['grades'] and candidate_sig['grades'] and source_sig['grades'] != candidate_sig['grades']:
        return False
    return True


def _material_match_score(source: str, product, require_name_match: bool = True) -> int:
    candidate_text = ' '.join([
        product.name or '',
        product.material or '',
        product.code or '',
        product.notes or '',
    ])
    source_sig = _strict_material_signature(source)
    candidate_sig = _strict_material_signature(candidate_text)
    if not _strict_signatures_match(source_sig, candidate_sig):
        return 0

    source_tokens = _material_tokens(source)
    if not source_tokens:
        return 0

    name_tokens = _material_tokens(product.name or '')
    detail_tokens = _material_tokens(candidate_text)
    if require_name_match and not (source_tokens & name_tokens):
        return 0
    candidate_tokens = name_tokens | detail_tokens
    common = source_tokens & candidate_tokens
    if not common:
        return 0

    dimension_tokens = {t for t in source_tokens if any(ch.isdigit() for ch in t)}
    material_tokens = source_tokens - dimension_tokens
    if dimension_tokens and not dimension_tokens.issubset(candidate_tokens):
        return 0
    if material_tokens and not material_tokens.intersection(candidate_tokens):
        return 0

    score = (len(source_tokens & name_tokens) * 14) + (len(source_tokens & detail_tokens) * 8)
    score += len(dimension_tokens & candidate_tokens) * 8

    code = (product.code or '').upper()
    if code.startswith('3TB-'):
        score -= 25
    return score


def _find_matching_raw_material(row: dict):
    """Find an existing raw material card for auto-generated material rows."""
    from app.models import Product

    if not row.get('is_auto_hammadde'):
        return None

    wanted_name = row.get('name') or ''
    wanted_unit = row.get('unit_type') or ''
    candidates = Product.query.filter(Product.is_active == True, Product.type == 'hammadde').all()

    scored = []
    for product in candidates:
        if (product.code or '').upper().startswith('3TB-'):
            continue
        score = _material_match_score(wanted_name, product)
        if score <= 0:
            continue
        if wanted_unit and product.unit_type == wanted_unit:
            score += 20
        elif wanted_unit and _units_compatible(product.unit_type, wanted_unit, row.get('weight_per_unit') or 0):
            score += 8
        elif wanted_unit and product.unit_type != wanted_unit:
            score -= 15
        scored.append((score, product))

    if not scored:
        return None

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_product = scored[0]
    return best_product if best_score >= 20 else None


def _find_costing_raw_material(row: dict, exclude_product_id: int = None):
    """Cost-only material match; allows legacy part-coded hammadde cards as a fallback."""
    from app.models import Product

    material_text = row.get('material') or ''
    name_text = row.get('name') or ''
    wanted_name = material_text if _is_priceable_raw_material_name(material_text) else name_text
    if name_text and material_text and material_text.lower() not in {'hammadde', 'hazır', 'hazir'}:
        wanted_name = f'{wanted_name} {name_text}'
    wanted_unit = row.get('unit_type') or ''
    candidates = Product.query.filter(Product.is_active == True, Product.type == 'hammadde').all()

    scored = []
    for product in candidates:
        if exclude_product_id and product.id == exclude_product_id:
            continue
        score = _material_match_score(wanted_name, product, require_name_match=False)
        if score <= 0:
            continue
        if wanted_unit and product.unit_type == wanted_unit:
            score += 25
        elif wanted_unit and _units_compatible(product.unit_type, wanted_unit, row.get('weight_per_unit') or 0):
            score += 12
        elif wanted_unit and product.unit_type != wanted_unit:
            score -= 20
        if product.unit_cost and product.unit_cost > 0:
            score += 30
        if not (product.code or '').upper().startswith('3TB-'):
            score += 10
        scored.append((score, product))

    if not scored:
        return None

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_product = scored[0]
    return best_product if best_score >= 20 else None


def _is_priceable_raw_material_name(name: str) -> bool:
    tokens = _material_tokens(name)
    if tokens & {'STANDART', 'HAZIR', 'DOVME', 'DÖVME'}:
        return False
    raw_markers = {
        'SAC', 'LAMA', 'BORU', 'PROFIL', 'TRANSMISYON', 'MIL',
        'CELIK', 'CEKME', 'C1040', 'SIYAH', 'ST37', 'ST44', 'SANAYI'
    }
    return bool(tokens & raw_markers)


def _is_ready_purchase_text(value: str) -> bool:
    text = _ascii_upper(value)
    return any(token in text for token in ('HAZIR', 'STANDART PARCA', 'STANDART'))


def _c(v) -> str:
    """None → '' temizleyici."""
    return '' if v is None else str(v).strip()


def _float(v, default=1.0) -> float:
    try:
        return float(str(v).replace(',', '.').strip())
    except Exception:
        return default


_NUM_RE = re.compile(r'^\s*\d+(\.\d+)*\.?\s*$')    # "1.", "1.1.", "1.1.2", "10.5.3."
_SEC_RE = re.compile(r'^\s*(\d+)\.\s+\S')          # "1. KAFA", "2. GÖVDE"


def _is_num(v: str) -> bool:
    """Noktalı BOM hiyerarşi numaraları: '1.', '1.1.', '1.1.2.' — en az bir nokta gerekli.
    Sıralı integer sayaçlarını (1, 2, 3...) BOM numarası olarak algılamamak için."""
    return '.' in v and bool(_NUM_RE.match(v))


def _normalize_num(v: str) -> str:
    """Son nokta yoksa ekle: "1.1.2" → "1.1.2." """
    s = v.strip()
    return s if s.endswith('.') else s + '.'


def _calc_level(num: str) -> int:
    """
    "1."       → 1
    "1.1."     → 2
    "1.1.1."   → 3
    """
    return len([p for p in num.split('.') if p.strip()])


def _unit_str(v: str) -> str:
    """Excel birim metnini normalleştirir.
    Desteklenen tipler: adet, kg, gr, ton, metre, cm, mm, m², m³, litre, paket, rulo, takim
    """
    if not v:
        return 'adet'
    ui = v.strip().lower()
    if not ui or ui in ('nan', 'none'):
        return 'adet'
    # Ağırlık
    if 'ton' in ui and 'kilo' not in ui:
        return 'ton'
    if 'kilogram' in ui or ui in ('kg', 'kilo'):
        return 'kg'
    if 'gram' in ui and 'kilo' not in ui:
        return 'gr'
    # Alan / hacim
    if 'metre kare' in ui or 'm2' in ui or 'm²' in ui or 'metrekare' in ui:
        return 'm²'
    if 'metre küp' in ui or 'm3' in ui or 'm³' in ui or 'metreküp' in ui:
        return 'm³'
    # Uzunluk
    if 'santimetre' in ui or ui in ('cm', 'santim'):
        return 'cm'
    if 'milimetre' in ui or ui == 'mm':
        return 'mm'
    if 'metre' in ui or ui in ('mt', 'm'):
        return 'metre'
    # Diğer
    if 'litre' in ui or ui in ('lt', 'l'):
        return 'litre'
    if 'rulo' in ui:
        return 'rulo'
    if 'paket' in ui or ui in ('pkt', 'pk'):
        return 'paket'
    if 'takım' in ui or 'takim' in ui or ui == 'set':
        return 'takım'
    if 'kutu' in ui:
        return 'kutu'
    # Bilinmeyeni olduğu gibi küçük harfle döndür
    return ui


def _units_compatible(product_unit: str, row_unit: str, weight_per_unit: float = 0.0) -> bool:
    product_unit = (product_unit or '').lower()
    row_unit = (row_unit or '').lower()
    if product_unit == row_unit:
        return True
    if weight_per_unit and product_unit in {'kg', 'gr', 'ton'} and row_unit in {'metre', 'adet'}:
        return True
    if weight_per_unit and product_unit == 'metre' and row_unit in {'kg', 'gr', 'ton'}:
        return True
    return False


def _cost_quantity_for_unit(product_unit: str, row_unit: str, quantity: float, piece_count: float = 1.0,
                            weight_per_unit: float = 0.0) -> float:
    """Convert BOM quantity to the Product card's pricing unit when possible."""
    product_unit = (product_unit or '').lower()
    row_unit = (row_unit or '').lower()
    qty = float(quantity or 0)
    pieces = float(piece_count or 1)
    weight = float(weight_per_unit or 0)

    if product_unit == row_unit:
        return qty
    if not weight:
        return qty
    if product_unit == 'kg' and row_unit == 'metre':
        return qty * weight
    if product_unit == 'kg' and row_unit == 'adet':
        return pieces * weight
    if product_unit == 'gr' and row_unit == 'metre':
        return qty * weight * 1000
    if product_unit == 'gr' and row_unit == 'adet':
        return pieces * weight * 1000
    if product_unit == 'ton' and row_unit == 'metre':
        return (qty * weight) / 1000
    if product_unit == 'ton' and row_unit == 'adet':
        return (pieces * weight) / 1000
    if product_unit == 'metre' and row_unit == 'kg':
        return qty / weight
    if product_unit == 'metre' and row_unit == 'gr':
        return (qty / 1000) / weight
    if product_unit == 'metre' and row_unit == 'ton':
        return (qty * 1000) / weight
    return qty


def _cost_basis_quantity(quantity: float, quantity_net: float = None) -> float:
    """Use net (firesiz) quantity for costing when available."""
    net = float(quantity_net or 0)
    if net > 0:
        return net
    return float(quantity or 0)


def _should_cost_by_weight(material_text: str, row_unit: str, weight_per_unit: float = 0.0) -> bool:
    """Sac/lama are priced by kg when BOM provides kg-per-unit data."""
    if not weight_per_unit:
        return False
    if (row_unit or '').lower() not in {'metre', 'mt', 'adet'}:
        return False
    signature = _strict_material_signature(material_text or '')
    return bool(signature and signature.get('family') in {'SAC', 'LAMA'})


def _weight_cost_quantity(quantity: float, weight_per_unit: float = 0.0) -> float:
    return float(quantity or 0) * float(weight_per_unit or 0)


# ---------------------------------------------------------------------------
# FORMAT TESPİTİ
# ---------------------------------------------------------------------------

def _detect_format(ws) -> str:
    """
    İlk 15 dolu satırı inceleyerek formatı belirle.
    ÖNCELİK 1: Eğer noktalı BOM numaraları (1.1.) varsa -> FORMAT A ('numbered')
    ÖNCELİK 2: Fireli/Firesiz başlıkları varsa -> FORMAT C ('formatted_bom')
    ÖNCELİK 3: Aksi halde Format B ('indented')
    """
    num_hits = 0
    for row in ws.iter_rows(max_row=20, values_only=True):
        vals = [_c(v) for v in row]
        if not any(vals):
            continue
        # Tüm hücreleri tara — BOM numarası ilk kolonda olmayabilir
        for v in vals:
            if v and (_is_num(v) or _SEC_RE.match(v)):
                num_hits += 1
                break
    
    if num_hits >= 2:
        return 'numbered'

    # FORMAT C tespiti: ÇOK KESİN — aynı satırda hem "fireli" hem "firesiz"
    for row in ws.iter_rows(max_row=3, values_only=True):
        vals = [_c(v).lower() for v in row if v]
        if len(vals) < 5:
            continue
        has_fireli  = any('fireli' in v for v in vals)
        has_firesiz = any('firesiz' in v for v in vals)
        has_metre_or_parca = any(('metre' in v or 'parça kodu' in v) for v in vals)
        if has_fireli and has_firesiz and has_metre_or_parca:
            return 'formatted_bom'

    return 'indented'


# ---------------------------------------------------------------------------
# FORMAT A — Numaralandırma Kolonu
# ---------------------------------------------------------------------------

def _parse_numbered(ws, override_root_name=None) -> tuple[list[dict], list[dict]]:
    rows   = []
    errors = []
    stack  = {}  # level → current num

    # ROW 1'İ root olarak kaydet (Ana ürün adı ilk dolu hücrede)
    first = list(ws.iter_rows(max_row=1, values_only=True))[0]
    root_name = ''
    for v in first:
        s = _c(v)
        # Tam sayı (0, 1, 2...), BOM numarası (1., 1.1.) ve bölüm başlığı (1. KAFA) dışındaki ilk değer ürn adı
        if s and not _is_num(s) and not _SEC_RE.match(s) and not re.match(r'^\d+$', s):
            root_name = s
            break
    if not root_name:
        root_name = 'ANA ÜRÜN'
    if override_root_name:
        root_name = override_root_name

    rows.append({'num': '0.', 'level': 0, 'name': root_name, 'code': '',
                 'material': '', 'quantity': 1.0, 'unit_type': 'adet',
                 'quantity_net': 1.0, 'piece_count': 1.0, 'weight_per_unit': 0.0, 'weight_unit': '',
                 'parent_num': None, 'excel_row': 1})
    stack[0] = '0.'

    # Başlık satırları — "Numara", "Adlandırması" gibi metinler içeren satırları bul/atla
    _HEADER_WORDS = {'numara', 'adlandırma', 'parca', 'parça', 'malzeme', 'kod',
                     'adet', 'açıklama', 'description', 'item', 'qty'}

    _INT_RE = re.compile(r'^\s*\d+\s*$')   # Sadece tam sayı: "0", "1", "123" — BOM numarası değil

    def _is_header_row(rv: list[str]) -> bool:
        non_empty = [v.lower() for v in rv if v and not re.match(r'^\d+$', v)]
        if not non_empty:
            return False
        return any(any(h in w for h in _HEADER_WORDS) for w in non_empty)

    col_map = {}
    col_map_found = False

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rv_raw = list(row_vals)
        rv = [_c(v) for v in rv_raw]
        if not any(rv):
            continue

        rv_lower = [str(x).lower().strip() for x in rv]
        if not col_map_found and any('numara' in x for x in rv_lower) and any('adlandır' in x for x in rv_lower):
            col_map_found = True
            for c, val in enumerate(rv_lower):
                if 'numara' in val: col_map['num'] = c
                elif 'adlandır' in val and 'name' not in col_map: col_map['name'] = c
                elif 'cinsi' in val: col_map['type'] = c
                elif 'kodu' in val: col_map['code'] = c
                elif 'özellik' in val or 'ozelligi' in val or 'özelliği' in val: col_map['spec'] = c
                elif 'fireli metre' in val: col_map['fireliM'] = c
                elif 'firesiz metre' in val: col_map['firesizM'] = c
                elif 'fireli ağırlık' in val or 'fireli agirlik' in val: col_map['fireliA'] = c
                elif 'firesiz ağırlık' in val or 'firesiz agirlik' in val: col_map['firesizA'] = c
                elif val == 'adet' or 'toplam adet' in val: col_map['qty'] = c
                elif 'fire miktarı' in val or 'fire miktari' in val or 'fire' in val: col_map['fireM'] = c
            continue

        # BOM numarası içeren ilk hücreyi bul (noktalı format aranır)
        first_val = ''
        first_col = -1
        
        if col_map_found and 'num' in col_map and col_map['num'] < len(rv):
            cv = rv[col_map['num']]
            if cv and (_is_num(cv) or _SEC_RE.match(cv)):
                first_val = cv
                first_col = col_map['num']

        if first_col == -1:
            for ci, v in enumerate(rv):
                if v and _is_num(v):
                    first_val = v
                    first_col = ci
                    break

        if first_col == -1:
            for ci, v in enumerate(rv):
                if v:
                    first_val = v
                    first_col = ci
                    break

        # Bölüm başlığı: "1. KAFA", "2. GÖVDE" — bunları L1 düğüm yap
        sec_m = _SEC_RE.match(first_val)
        if sec_m:
            cnt1 = len([r for r in rows if r['level'] == 1]) + 1
            num = f'{cnt1}.'
            group_name = re.sub(r'^\s*\d+\.\s*', '', first_val).strip() or first_val
            parent_num = stack.get(0)
            stack[1] = num
            for k in [l for l in list(stack) if l > 1]: del stack[k]
            rows.append({'num': num, 'level': 1, 'name': group_name, 'code': '',
                         'material': '', 'quantity': 1.0, 'unit_type': 'adet',
                         'quantity_net': 1.0, 'piece_count': 1.0, 'weight_per_unit': 0.0, 'weight_unit': '',
                         'parent_num': parent_num, 'excel_row': row_idx})
            continue

        # Başlık satırı → atla
        if _is_header_row(rv) and not _is_num(first_val):
            continue

        # Numara hücresi — noktalı BOM numarası olmalı
        if not _is_num(first_val):
            continue  # değer yok veya tanınmıyor

        num = _normalize_num(first_val)
        level = _calc_level(num)

        name = ''
        code = ''
        material = ''
        qty_fireli  = 0.0
        qty_firesiz = 0.0
        unit        = 'adet'
        w_val  = 0.0
        w_unit = ''

        if col_map_found and 'name' in col_map:
            name = _c(rv[col_map['name']]) if col_map['name'] < len(rv) else ''
            typ = _c(rv[col_map['type']]) if 'type' in col_map and col_map['type'] < len(rv) else ''
            spc = _c(rv[col_map['spec']]) if 'spec' in col_map and col_map['spec'] < len(rv) else ''
            # Ölçü (typ = Malzeme Cinsi: Ø76×5) ve Özellik (spc: Sanayi Borusu) birleştir => "Sanayi Borusu Ø76x5"
            if typ and spc:
                material = f"{spc} {typ}".strip()
            else:
                material = typ or spc
            code = _c(rv[col_map['code']]) if 'code' in col_map and col_map['code'] < len(rv) else ''
            
            e_val = _float(rv_raw[col_map['fireliM']], 0.0) if 'fireliM' in col_map and col_map['fireliM'] < len(rv_raw) else 0.0
            f_val = _float(rv_raw[col_map['firesizM']], 0.0) if 'firesizM' in col_map and col_map['firesizM'] < len(rv_raw) else 0.0
            g_val = _float(rv_raw[col_map['fireliA']], 0.0) if 'fireliA' in col_map and col_map['fireliA'] < len(rv_raw) else 0.0
            h_val = _float(rv_raw[col_map['firesizA']], 0.0) if 'firesizA' in col_map and col_map['firesizA'] < len(rv_raw) else 0.0
            i_val = _float(rv_raw[col_map['qty']], 1.0) if 'qty' in col_map and col_map['qty'] < len(rv_raw) else 1.0
            
            if e_val > 0:
                qty_fireli  = e_val
                qty_firesiz = f_val if f_val > 0 else e_val
                unit        = 'metre'
                w_val  = round(g_val / e_val, 4) if g_val > 0 and e_val > 0 else 0.0
                w_unit = 'kg'
                pc_val = i_val
            elif g_val > 0:
                qty_fireli  = g_val
                qty_firesiz = h_val if h_val > 0 else g_val
                unit        = 'kg'
                w_val  = 0.0
                w_unit = 'kg'
                pc_val = i_val
            else:
                qty_fireli  = i_val
                qty_firesiz = i_val
                pc_val = i_val
                unit        = 'adet'
                w_val  = 0.0
                w_unit = ''
                
        else:
            remaining = [(ci, v) for ci, v in enumerate(rv) if ci > first_col and v]
            if remaining:
                name = remaining[0][1]
            candidate_material = _c(rv[first_col + 2]) if len(rv) > first_col + 2 else ''
            candidate_code = _c(rv[first_col + 3]) if len(rv) > first_col + 3 else ''
            code = candidate_code
            material = candidate_material

            if not material and not code and len(remaining) >= 2:
                candidate = remaining[1][1]
                if re.match(r'^[A-Za-z0-9\-]{4,}$', candidate.replace(' ', '')) and ('-' in candidate or any(c.isdigit() for c in candidate)):
                    code = candidate
                    if len(remaining) >= 3:
                        material = remaining[2][1]
                else:
                    material = candidate
                    if len(remaining) >= 3:
                        cand3 = remaining[2][1]
                        if '-' in cand3 or any(d.isdigit() for d in cand3):
                            code = cand3

            qty_raw = None
            for _, v in reversed(remaining):
                f = _float(v, default=-1)
                if 0 < f < 100000:
                    qty_raw = f
                    break
            unit_raw = ''
            for _, v in remaining:
                if v.lower() in ('adet', 'kg', 'kilogram', 'metre', 'mt', 'm2'):
                    unit_raw = v
                    break

            qty_fireli  = qty_raw if qty_raw is not None else 1.0
            qty_firesiz = qty_fireli
            unit = _unit_str(unit_raw)
            pc_val = 1.0  # Default piece count
            w_val = 0.0
            w_unit = ''

        if not name:
            errors.append({'row': row_idx, 'error': f'{num} için ad bulunamadı — atlandı'})
            continue

        for k in [l for l in list(stack) if l > level]: del stack[k]
        parent_num = stack.get(level - 1)
        stack[level] = num

        rows.append({'num': num, 'level': level, 'name': name, 'code': code,
                     'material': material, 'quantity': qty_fireli, 'unit_type': unit,
                     'quantity_net': qty_firesiz, 'weight_per_unit': w_val, 'weight_unit': w_unit,
                         'piece_count': pc_val,
                     'parent_num': parent_num, 'excel_row': row_idx})

    return rows, errors


# ---------------------------------------------------------------------------
# FORMAT B — Kolon Girintisi (ÇELMAK ürün ağacı)
# ---------------------------------------------------------------------------
# Sütun düzeni (GERGEÇ Excel yapısı):
#   A (0): Ana ürün adı   B (1): L1 grup         C (2): L2 alt grup
#   D (3): Parça adı     E (4): Malzeme/Özellik
#   F (5): Birim ağırlık  G (6): Ağırlık birimi (Kilogram)
#   H (7): Firesiz adet  I (8): Birim (Adet)
#   J (9): Toplam ağırlık  K (10): Ağırlık birimi
#   L (11): Fireli adet  M (12): Birim (Adet)
# ---------------------------------------------------------------------------

def _parse_indented(ws, override_root_name=None) -> tuple[list[dict], list[dict]]:
    rows   = []
    errors = []
    stack  = {0: None}
    cnt    = {1: 0, 2: 0, 3: 0}

    # İlk satırdan ana ürün adını al (Col A)
    first = list(ws.iter_rows(max_row=1, values_only=True))[0]
    root_name = _c(first[0]) if first else 'ANA ÜRÜN'
    if not root_name:
        root_name = 'ANA ÜRÜN'
    if override_root_name:
        root_name = override_root_name

    rows.append({'num': '0.', 'level': 0, 'name': root_name, 'code': '',
                 'material': '', 'quantity': 1.0, 'unit_type': 'adet',
                 'quantity_net': 1.0, 'piece_count': 1.0, 'weight_per_unit': 0.0, 'weight_unit': '',
                 'parent_num': None, 'excel_row': 1})
    stack[0] = '0.'

    last_l1 = None
    last_l2 = None

    def emit(level_: int, name_: str,
             qty_fireli: float, qty_firesiz: float,
             unit_: str, mat_: str,
             weight_val: float, weight_u: str, ridx: int):
        nonlocal last_l1, last_l2
        if level_ == 1:
            cnt[1] += 1; cnt[2] = 0; cnt[3] = 0
            num_ = f'{cnt[1]}.'; last_l1 = name_; last_l2 = None
        elif level_ == 2:
            cnt[2] += 1; cnt[3] = 0
            num_ = f'{cnt[1]}.{cnt[2]}.'; last_l2 = name_
        else:
            cnt[3] += 1
            num_ = (f'{cnt[1]}.{cnt[2]}.{cnt[3]}.' if cnt[2] > 0
                    else f'{cnt[1]}.0.{cnt[3]}.')

        # En yakın mütıfar eşleşitir — boşluk olsa bile köke fallback yap
        parent_ = None
        for lv in range(level_ - 1, -1, -1):
            p = stack.get(lv)
            if p is not None:
                parent_ = p
                break
        stack[level_] = num_
        for k in [l for l in list(stack) if l > level_]: del stack[k]

        rows.append({
            'num': num_, 'level': level_, 'name': name_, 'code': '',
            'material': mat_, 'quantity': qty_fireli, 'unit_type': unit_,
            'quantity_net': qty_firesiz,
            'weight_per_unit': weight_val, 'weight_unit': weight_u,
            'parent_num': parent_, 'excel_row': ridx
        })

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        rv = list(row_vals)
        col_b = _c(rv[1])  if len(rv) > 1  else ''   # L1 grup
        col_c = _c(rv[2])  if len(rv) > 2  else ''   # L2 alt grup
        col_d = _c(rv[3])  if len(rv) > 3  else ''   # Parça adı
        col_e = _c(rv[4])  if len(rv) > 4  else ''   # Malzeme / Özellik / Boyut
        col_f = rv[5]      if len(rv) > 5  else None  # Birim ağırlık değeri
        col_g = _c(rv[6])  if len(rv) > 6  else ''   # Ağırlık birimi (Kilogram)
        col_h = rv[7]      if len(rv) > 7  else None  # Firesiz adet
        col_i = _c(rv[8])  if len(rv) > 8  else ''   # Firesiz birimi (Adet/Metre/Kg)
        col_l = rv[11]     if len(rv) > 11 else None  # Fireli (toplam) adet
        col_m = _c(rv[12]) if len(rv) > 12 else ''   # Fireli birimi (ASIL birim — öncelikli)

        if not col_b and not col_c and not col_d:
            continue

        # ------- Miktar: fireli (col L) öncelikli, yoksa firesiz, yoksa 1 -------
        qty_fireli  = _float(col_l, default=0.0) if col_l is not None else 0.0
        qty_firesiz = _float(col_h, default=0.0) if col_h is not None else 0.0
        qty_use     = (qty_fireli  if qty_fireli  > 0 else
                       qty_firesiz if qty_firesiz > 0 else 1.0)

        # ------- Birim: col M (fireli birimi) öncelikli, yoksa col I (firesiz birimi) -------
        # Col M ve Col I farklı olabilir ama çoğunlukla aynıdır.
        # Col M = fireli adede ait birim → asil kullanım birimi
        raw_unit = col_m if (col_m and col_m.lower() not in ('nan', 'none', '')) else col_i
        unit_str = _unit_str(raw_unit)

        # ------- Ağırlık bilgisi -------
        # Col F = birim ağırlık (her zaman kg cinsindendir)
        # Col G = ağırlık birimi (genellikle 'Kilogram')
        # Birim zaten 'kg' ise (ağırlık bazlı ölçüm), col F ayrıca saklanmaz — miktarın kendisi ağırlık
        weight_val = _float(col_f, default=0.0) if col_f is not None else 0.0
        col_g_lo   = col_g.lower()
        weight_u   = ('kg' if ('kilo' in col_g_lo or col_g_lo in ('kg', ''))
                      else _unit_str(col_g))

        if row_idx == 1:
            if col_b and col_b != last_l1:
                emit(1, col_b, 1.0, 1.0, 'adet', '', 0.0, '', row_idx)
            if col_c and col_c != last_l2:
                emit(2, col_c, 1.0, 1.0, 'adet', '', 0.0, '', row_idx)
            if col_d:
                emit(3, col_d, qty_use, qty_firesiz, unit_str, col_e,
                     weight_val, weight_u, row_idx)
            continue

        if col_b and col_b != last_l1:
            emit(1, col_b, 1.0, 1.0, 'adet', '', 0.0, '', row_idx)
        if col_c and col_c != last_l2:
            emit(2, col_c, 1.0, 1.0, 'adet', '', 0.0, '', row_idx)
        if col_d:
            emit(3, col_d, qty_use, qty_firesiz, unit_str, col_e,
                 weight_val, weight_u, row_idx)

    return rows, errors


# ---------------------------------------------------------------------------
# FORMAT C — Başlık Satırlı BOM (Parça Kodu, Fireli/Firesiz Metre & Ağırlık)
# ---------------------------------------------------------------------------
# Sütun düzeni (başlık satırından sonra):
#   A (0): Adlandırılması (parça/montaj adı)
#   B (1): Malzeme Cinsi (e.g., "185x75x4", "4 mm", "Montaj")
#   C (2): Parça Kodu    (e.g., "3TB-P0201-099-50")
#   D (3): Malzeme Özelliği (e.g., "Profil", "Siyah Sac (St3237)")
#   E (4): Toplam Fireli Metre
#   F (5): Toplam Firesiz Metre
#   G (6): Toplam Fireli Ağırlık (kg)
#   H (7): Toplam Firesiz Ağırlık (kg)
#   I (8): Toplam Adet
#   J (9): Toplam Fire
#
# Montaj (assembly) satırı tespiti:
#   1. Sarı/altın renk dolgu (Excel cell fill)
#   2. Fallback: Col B "Montaj" içeriyorsa
#   3. Fallback: Tüm E-H değerleri sıfır ama Col I > 0
# ---------------------------------------------------------------------------

def _cell_is_yellow(cell) -> bool:
    """Hücre dolgusu sarı/altın tonundaysa True."""
    try:
        fill = cell.fill
        if fill.fill_type == 'solid':
            rgb = fill.fgColor.rgb  # ARGB: 'FFRRGGBB'
            if len(rgb) == 8:
                r = int(rgb[2:4], 16)
                g = int(rgb[4:6], 16)
                b = int(rgb[6:8], 16)
                # Sarı/altın/turuncu: yüksek R, orta-yüksek G, düşük B
                if r > 180 and g > 120 and b < 80:
                    return True
    except Exception:
        pass
    return False


def _is_format_c_header(sv: list[str]) -> bool:
    """Bir satırın FORMAT C başlık satırı olup olmadığını kontrol eder."""
    if len([v for v in sv if v]) < 5:
        return False
    vals = [v.lower() for v in sv if v]
    has_fireli  = any('fireli' in v for v in vals)
    has_firesiz = any('firesiz' in v for v in vals)
    has_metre_or_parca = any(('metre' in v or 'parça kodu' in v) for v in vals)
    return has_fireli and has_firesiz and has_metre_or_parca


def _parse_format_c(ws, override_root_name=None) -> tuple[list[dict], list[dict]]:
    rows   = []
    errors = []

    # İlk 3 satırı tara: başlık satırı satır numarasını bul, ürün adını tespit et
    header_row_idx = None
    root_name      = 'ANA ÜRÜN'
    for i, row_vals in enumerate(ws.iter_rows(max_row=3, values_only=True), start=1):
        sv = [_c(v) for v in row_vals]
        if _is_format_c_header(sv):
            header_row_idx = i
            break
        else:
            # Başlık değil → ürün adı adayı
            non_empty = [v for v in sv if v]
            if non_empty:
                root_name = non_empty[0]
                
    if override_root_name:
        root_name = override_root_name

    rows.append({'num': '0.', 'level': 0, 'name': root_name, 'code': '',
                 'material': '', 'quantity': 1.0, 'unit_type': 'adet',
                 'quantity_net': 1.0, 'piece_count': 1.0, 'weight_per_unit': 0.0, 'weight_unit': '',
                 'parent_num': None, 'excel_row': 0})

    l1_cnt = 0
    l2_cnt = 0
    current_l1_num = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        rv   = [cell.value for cell in row]
        sv   = [_c(v) for v in rv]

        # Başlık veya başlık öncesi satırları atla
        if header_row_idx is not None and row_idx <= header_row_idx:
            continue
        # Başlık satırı tespit edilemedi ise: ilk karşılaşılan header satırını atla
        if header_row_idx is None and _is_format_c_header(sv):
            header_row_idx = row_idx
            continue

        if not any(sv):
            continue  # boş satır

        col_a = sv[0]  if len(sv) > 0 else ''   # Ad
        col_b = sv[1]  if len(sv) > 1 else ''   # Malzeme Cinsi
        col_c = sv[2]  if len(sv) > 2 else ''   # Parça Kodu
        col_d = sv[3]  if len(sv) > 3 else ''   # Malzeme Özelliği
        e_val = _float(rv[4], 0.0) if len(rv) > 4 and rv[4] is not None else 0.0  # Fireli Metre
        f_val = _float(rv[5], 0.0) if len(rv) > 5 and rv[5] is not None else 0.0  # Firesiz Metre
        g_val = _float(rv[6], 0.0) if len(rv) > 6 and rv[6] is not None else 0.0  # Fireli Ağırlık
        h_val = _float(rv[7], 0.0) if len(rv) > 7 and rv[7] is not None else 0.0  # Firesiz Ağırlık
        i_val = _float(rv[8], 1.0) if len(rv) > 8 and rv[8] is not None else 1.0  # Adet

        if not col_a:
            continue

        # ——— Montaj (assembly) satırı mı? ———
        # Güvenilir göstergeler: sarı hücre dolgusu VEYA col_b = "Montaj"
        # NOT: E-H=0 ve I>0 koşulu adet bazlı bileşenleri de yakalar → kullanılmaz
        is_assembly = _cell_is_yellow(row[0])
        if not is_assembly:
            col_b_low = col_b.lower()
            if col_b_low in ('montaj', 'assembly', 'alt montaj', 'submontaj', 'montage'):
                is_assembly = True

        if is_assembly:
            l1_cnt += 1
            l2_cnt = 0
            current_l1_num = f'{l1_cnt}.'
            adet = i_val if i_val > 0 else 1.0
            rows.append({
                'num': current_l1_num, 'level': 1,
                'name': col_a, 'code': col_c,
                'material': col_b, 'quantity': adet, 'unit_type': 'adet',
                'quantity_net': adet,
                'weight_per_unit': 0.0, 'weight_unit': '',
                'parent_num': '0.', 'excel_row': row_idx
            })
        else:
            # ——— Bileşen satırı ———
            if current_l1_num is None:
                # Henüz montaj başlığı yok → genel grup oluştur
                l1_cnt += 1
                l2_cnt = 0
                current_l1_num = f'{l1_cnt}.'
                rows.append({
                    'num': current_l1_num, 'level': 1,
                    'name': 'Genel', 'code': '',
                    'material': '', 'quantity': 1.0, 'unit_type': 'adet',
                    'quantity_net': 1.0, 'piece_count': 1.0, 'weight_per_unit': 0.0, 'weight_unit': '',
                    'parent_num': '0.', 'excel_row': row_idx
                })

            l2_cnt += 1
            num = f'{l1_cnt}.{l2_cnt}.'

            # ——— Birim ve miktar tespiti ———
            if e_val > 0:
                # Metre bazlı ölçüm
                qty_fireli  = e_val
                qty_firesiz = f_val if f_val > 0 else e_val
                unit        = 'metre'
                # Birim ağırlık = toplam ağırlık / metre
                w_val  = round(g_val / e_val, 4) if g_val > 0 and e_val > 0 else 0.0
                w_unit = 'kg'
                pc_val = i_val
            elif g_val > 0:
                # Ağırlık bazlı ölçüm (sac, levha vb.)
                qty_fireli  = g_val
                qty_firesiz = h_val if h_val > 0 else g_val
                unit        = 'kg'
                w_val  = 0.0
                w_unit = 'kg'
                pc_val = i_val
            else:
                # Adet bazlı (aksesuar, bağlantı elemanı vb.)
                qty_fireli  = i_val
                qty_firesiz = i_val
                pc_val = i_val
                unit        = 'adet'
                w_val  = 0.0
                w_unit = ''

            # material: Özel format (Özellik + Cins) => Örn: "Sanayi Borusu Ø76x5 mm"
            if col_b and col_d and col_b.lower() not in ('montaj', 'assembly', 'alt montaj', 'submontaj', 'montage'):
                mat = f"{col_d} {col_b}".strip()
            else:
                mat = col_d if col_d else col_b

            rows.append({
                'num': num, 'level': 2,
                'name': col_a, 'code': col_c,
                'material': mat, 'quantity': qty_fireli, 'unit_type': unit,
                'quantity_net': qty_firesiz,
                'weight_per_unit': w_val, 'weight_unit': w_unit,
                         'piece_count': pc_val,
                'parent_num': current_l1_num, 'excel_row': row_idx
            })

    return rows, errors


# ---------------------------------------------------------------------------
# ANA GİRİŞ NOKTASI
# ---------------------------------------------------------------------------

def parse_bom_excel_v2(file_stream, override_root_name=None) -> tuple[list[dict], list[dict]]:
    """
    ÇELMAK ürün ağacı Excel dosyasını parse eder.
    Format otomatik tespit edilir (numara kolonlu veya kolon girintili).
    """
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active
    except Exception as exc:
        return [], [{'row': 0, 'error': f'Dosya okuma hatası: {exc}'}]

    fmt = _detect_format(ws)

    if fmt == 'formatted_bom':
        rows, errors = _parse_format_c(ws, override_root_name)
    elif fmt == 'numbered':
        rows, errors = _parse_numbered(ws, override_root_name)
    else:
        rows, errors = _parse_indented(ws, override_root_name)

    if len(rows) <= 1:
        return [], [{'row': 0, 'error':
            f'Excel\'den hiç parça çıkarılamadı (format={fmt}). '
            'Numara kolonunu veya kolon yapısını kontrol edin.'}]

    # --- YarıMamül ve Otomatik Hammadde Dönüşümü ---
    transformed_rows = []
    for r in rows:
        mat_str = str(r.get('material', '')).strip()
        c_prefix = str(r.get('code') or '')[:3]
        
        is_assembly = mat_str.lower() in ('montaj', 'assembly', 'alt montaj', 'submontaj', 'montage')
        is_standard = c_prefix in STANDARD_PREFIXES or mat_str.lower() == 'standart parça'
        is_ready_purchase = _is_ready_purchase_text(mat_str)
        
        # Eğer bir montaj/seviye 0 değilse, standart parça değilse ve malzemesi varsa:
        needs_child = not is_assembly and not is_standard and not is_ready_purchase and bool(mat_str) and r['level'] > 0
        
        if is_standard or is_ready_purchase:
            # Standart Parçaların birimi her zaman "adet" olmalıdır
            parent_qty = r.get('piece_count', 1.0)
            r['quantity'] = parent_qty
            r['quantity_net'] = parent_qty
            r['unit_type'] = 'adet'
            if is_ready_purchase and not is_standard:
                r['is_ready_purchase'] = True
            transformed_rows.append(r)

        elif needs_child:
            # Yarı Mamul'ün kendi adedi
            parent_qty = r.get('piece_count', 1.0)
            child_qty_fireli = r.get('quantity', parent_qty)
            child_qty_firesiz = r.get('quantity_net', parent_qty)
            child_unit = r.get('unit_type', 'adet')
            
            # Yarı Mamul'ü Adet olarak sabitle
            r['quantity'] = parent_qty
            r['quantity_net'] = parent_qty
            r['unit_type'] = 'adet'
            
            transformed_rows.append(r)
            
            # Hammadde child node oluştur
            child = {
                'num': r['num'] + '1.',
                'level': r['level'] + 1,
                'name': mat_str,  # Malzeme Cinsi + Özelliği
                'code': '',       # Hammaddenin kendi kodu yok, ürün koduna geçmesin
                'material': 'Hammadde',
                'quantity': child_qty_fireli,
                'quantity_net': child_qty_firesiz,
                'unit_type': child_unit,
                'piece_count': 1.0,
                'weight_per_unit': r.get('weight_per_unit', 0.0),
                'weight_unit': r.get('weight_unit', ''),
                'parent_num': r['num'],
                'excel_row': r.get('excel_row', 0),
                'is_auto_hammadde': True
            }
            transformed_rows.append(child)
        else:
            transformed_rows.append(r)

    return transformed_rows, errors


# ---------------------------------------------------------------------------
# BOM Analiz ve Önizleme
# ---------------------------------------------------------------------------

def analyze_bom_for_import(parsed_rows: list[dict], category_id: int = None) -> dict:
    """
    Excel'den parse edilmiş BOM satırlarını analiz eder.
    
    Returns:
        {
            'new_products': [...],      # Yeni eklenecek ürünler
            'existing_products': [...],  # Zaten mevcut ürünler
            'conflicts': [...],          # Çakışma/uyarı gerektiren durumlar
            'stats': {...}               # Özet istatistikler
        }
    """
    from app.models import Product, BomItem
    
    new_products = []
    existing_products = []
    conflicts = []
    matched_materials = []
    missing_materials = []
    
    # Excel dosyasındaki tekrarları (aynı isme sahip aynı parçalar) yakalamak için
    seen_in_excel = {}
    
    for row in parsed_rows:
        code_prefix = str(row.get('code') or '')[:3]
        if code_prefix in STANDARD_PREFIXES or str(row.get('material', '')).lower() == 'standart parça':
            item_type = 'standart_parca'
        elif row['level'] == 0:
            item_type = 'mamul'
        elif row.get('is_ready_purchase') or _is_ready_purchase_text(row.get('material') or ''):
            item_type = 'hazir_parca'
        elif row.get('is_auto_hammadde'): 
            item_type = 'hammadde'
        else:
            item_type = 'yarimamul'
        
        # Mevcut ürünü kontrol et (Adına ve Koda göre)
        excel_code = row.get('code') or ''
        product = Product.query.filter_by(name=row['name']).filter(Product.code == excel_code if excel_code else True).first()
        if not product:
            # Eğer koduyla bulamadıysa ama ismiyle ve boş koduyla varsa falan diye sadece name ile tekrar bakalım, 
            # ancak bu sefer excel_code doluysa ve db_code dolu ve farklıysa BUNLAR FARKLI ÜRÜNDÜR, AYNI SAYMA
            potentials = Product.query.filter_by(name=row['name']).all()
            for p in potentials:
                p_code = p.code or ''
                if not excel_code or not p_code or excel_code == p_code:
                    product = p
                    break
        
        item = BomItem.query.filter_by(name=row['name']).filter(BomItem.code == excel_code if excel_code else True).first()
        if not item:
            potentials_item = BomItem.query.filter_by(name=row['name']).all()
            for i in potentials_item:
                i_code = i.code or ''
                if not excel_code or not i_code or excel_code == i_code:
                    item = i
                    break

        matched_raw_material = None
        if not product and row.get('is_auto_hammadde'):
            matched_raw_material = _find_matching_raw_material(row)
            if matched_raw_material:
                product = matched_raw_material
        
        entry = {
            'name': row['name'],
            'code': row.get('code') or '',
            'material': row.get('material') or '',
            'unit_type': row['unit_type'],
            'quantity': row['quantity'],
            'item_type': item_type,
            'level': row['level'],
            'is_auto_hammadde': bool(row.get('is_auto_hammadde')),
            'matched_by': 'material' if matched_raw_material else None
        }
        
        # Eğer bu isimli ürün veritabanında yoksa ama bu excel dosyasında daha önce gördüysek, 
        # onu yeni bir ürün olarak tekrar tekrar eklemek yerine "mevcut_excel_tekrarı" gibi değerlendireceğiz.
        
        if not product:
            if row.get('is_auto_hammadde'):
                if row['name'] not in seen_in_excel and _is_priceable_raw_material_name(row['name']):
                    base_code = _make_product_code(row['name'], row.get('code') or '')
                    entry['generated_code'] = _unique_product_code(base_code)
                    entry['material'] = row['name']
                    new_products.append(entry)
                    seen_in_excel[row['name']] = True
                elif row['name'] not in seen_in_excel:
                    entry['reason'] = 'Mevcut hammadde kartı bulunamadı'
                    missing_materials.append(entry)
                    seen_in_excel[row['name']] = True
                continue

            if row['name'] not in seen_in_excel:
                # Gerçekten yeni ürün
                base_code = _make_product_code(row['name'], row.get('code') or '')
                entry['generated_code'] = _unique_product_code(base_code)
                new_products.append(entry)
                seen_in_excel[row['name']] = True
        else:
            # Mevcut ürün - çakışma kontrolü
            entry['existing_code'] = product.code
            entry['existing_material'] = product.material or ''
            entry['existing_unit_type'] = product.unit_type
            entry['existing_type'] = product.type
            entry['current_stock'] = product.current_stock
            
            # Çakışma tespiti
            issues = []
            updates = []
            
            # Birim tipi farklıysa UYARI
            if not _units_compatible(product.unit_type, row['unit_type'], row.get('weight_per_unit') or 0):
                issues.append({
                    'type': 'unit_mismatch',
                    'message': f"Birim tipi farklı: Mevcut '{product.unit_type}' vs Excel '{row['unit_type']}'"
                })
            
            # Tip farklıysa UYARI
            if product.type != item_type:
                issues.append({
                    'type': 'type_mismatch',
                    'message': f"Ürün tipi farklı: Mevcut '{product.type}' vs Excel '{item_type}'"
                })
            
            # Malzeme bilgisi yoksa eklenecek
            if matched_raw_material:
                entry['matched_code'] = product.code
                entry['matched_name'] = product.name
            elif row.get('material') and not product.material:
                pass
            elif row.get('material') and product.material and product.material != row['material']:
                issues.append({
                    'type': 'material_mismatch',
                    'message': f"Malzeme farklı: Mevcut '{product.material}' vs Excel '{row['material']}'"
                })
            
            entry['issues'] = issues
            entry['updates'] = updates
            
            if issues:
                conflicts.append(entry)
            elif matched_raw_material:
                matched_materials.append(entry)
            else:
                existing_products.append(entry)
    
    stats = {
        'total': len(parsed_rows),
        'new': len(new_products),
        'existing': len(existing_products),
        'conflicts': len(conflicts),
        'will_update': sum(1 for p in existing_products if p.get('updates')),
        'matched_materials': len(matched_materials),
        'missing_materials': len(missing_materials)
    }
    
    return {
        'new_products': new_products,
        'existing_products': existing_products,
        'conflicts': conflicts,
        'matched_materials': matched_materials,
        'missing_materials': missing_materials,
        'raw_material_summary': _aggregate_raw_material_usage(parsed_rows),
        'stats': stats
    }


def _aggregate_raw_material_usage(rows: list[dict]) -> list[dict]:
    grouped = {}
    for row in rows:
        if not row.get('is_auto_hammadde') or not _is_priceable_raw_material_name(row.get('name') or ''):
            continue
        key = (_ascii_upper(row.get('name') or ''), row.get('unit_type') or 'adet')
        item = grouped.setdefault(key, {
            'name': row.get('name') or '',
            'unit_type': row.get('unit_type') or 'adet',
            'quantity': 0.0,
            'quantity_net': 0.0,
            'usage_count': 0,
            'generated_code': _make_product_code(row.get('name') or '', row.get('code') or ''),
        })
        item['quantity'] += float(row.get('quantity') or 0)
        item['quantity_net'] += float(row.get('quantity_net') or 0)
        item['usage_count'] += 1

    return sorted(grouped.values(), key=lambda x: (x['name'], x['unit_type']))


def _compare_key(row: dict) -> tuple:
    code = str(row.get('code') or '').strip().upper()
    if code:
        return ('code', code)
    return ('name', _ascii_upper(row.get('name') or ''), int(row.get('level') or 0))


def _indexed_compare_rows(rows: list[dict]) -> dict:
    seen = Counter()
    indexed = {}
    for row in rows:
        base_key = _compare_key(row)
        seen[base_key] += 1
        indexed[(base_key, seen[base_key])] = row
    return indexed


def _row_changed_fields(old_row: dict, new_row: dict) -> list[dict]:
    checks = [
        ('name', 'Ad'),
        ('quantity', 'Fireli miktar'),
        ('quantity_net', 'Firesiz miktar'),
        ('piece_count', 'Parça adedi'),
        ('unit_type', 'Birim'),
        ('material', 'Malzeme'),
    ]
    changes = []
    for field, label in checks:
        old_val = old_row.get(field)
        new_val = new_row.get(field)
        if field in {'quantity', 'quantity_net', 'piece_count'}:
            try:
                old_cmp = round(float(old_val or 0), 4)
                new_cmp = round(float(new_val or 0), 4)
            except Exception:
                old_cmp = old_val
                new_cmp = new_val
        else:
            old_cmp = str(old_val or '').strip()
            new_cmp = str(new_val or '').strip()
        if old_cmp != new_cmp:
            changes.append({'field': field, 'label': label, 'old': old_val or '-', 'new': new_val or '-'})
    return changes


def _existing_bom_rows(bom_id: int) -> list[dict]:
    from app.models import BomNode

    nodes = BomNode.query.filter_by(bom_id=bom_id).order_by(BomNode.num).all()
    rows = []
    for node in nodes:
        item = node.item
        product = item.product if item else None
        rows.append({
            'num': node.num,
            'level': node.level,
            'name': node.display_name or (item.name if item else ''),
            'code': item.code if item else '',
            'material': product.material if product else '',
            'quantity': float(node.quantity or 0),
            'quantity_net': float(node.quantity_net or 0) if node.quantity_net is not None else 0,
            'piece_count': float(node.piece_count or 1),
            'unit_type': node.unit_type or (item.unit_type if item else 'adet'),
        })
    return rows


def _find_product_for_row(row: dict, for_cost: bool = False):
    from app.models import Product

    excel_code = row.get('code') or ''
    product = Product.query.filter_by(name=row['name']).filter(Product.code == excel_code if excel_code else True).first()
    if not product:
        potentials = Product.query.filter_by(name=row['name']).all()
        for p in potentials:
            p_code = p.code or ''
            if not excel_code or not p_code or excel_code == p_code:
                product = p
                break
    if not product and row.get('is_auto_hammadde'):
        product = _find_matching_raw_material(row)
    if product and for_cost:
        costing_row = dict(row)
        if product.material:
            costing_row['material'] = product.material
        priced_product = _find_costing_raw_material(costing_row, exclude_product_id=product.id)
        product_code = (product.code or '').upper()
        priced_code = (priced_product.code or '').upper() if priced_product else ''
        should_use_priced = (
            priced_product
            and priced_product.unit_cost
            and priced_product.unit_cost > 0
            and (
                not (product.unit_cost and product.unit_cost > 0)
                or (product_code.startswith('3TB-') and not priced_code.startswith('3TB-'))
            )
        )
        if should_use_priced:
            product = priced_product
    if not product and for_cost and (
        row.get('is_auto_hammadde')
        or str(row.get('material') or '').lower() == 'hammadde'
        or _is_priceable_raw_material_name(row.get('material') or row.get('name') or '')
    ):
        product = _find_costing_raw_material(row)
    return product


def estimate_bom_rows_cost(rows: list[dict]) -> float:
    """Best-effort preview cost for parsed rows using existing Product cards."""
    parent_nums = {row.get('parent_num') for row in rows if row.get('parent_num')}
    total = 0.0
    for row in rows:
        if row.get('num') in parent_nums:
            continue
        product = _find_product_for_row(row, for_cost=True)
        if not product:
            continue
        unit_cost = float(product.unit_cost or 0)
        cost_basis_qty = _cost_basis_quantity(row.get('quantity') or 0, row.get('quantity_net'))
        if _should_cost_by_weight(row.get('material') or row.get('name') or '', row.get('unit_type'), row.get('weight_per_unit') or 0):
            qty = _weight_cost_quantity(cost_basis_qty, row.get('weight_per_unit') or 0)
        else:
            qty = _cost_quantity_for_unit(
                product.unit_type,
                row.get('unit_type'),
                cost_basis_qty,
                row.get('piece_count') or 1,
                row.get('weight_per_unit') or 0
            )
        total += unit_cost * qty
    return round(total, 2)


def compare_bom_update(existing_bom_id: int, new_rows: list[dict], db) -> dict:
    """Compare an existing BOM with newly parsed rows before creating a revision."""
    old_rows = _existing_bom_rows(existing_bom_id)
    old_index = _indexed_compare_rows(old_rows)
    new_index = _indexed_compare_rows(new_rows)
    new_children_by_parent = {}
    for row in new_rows:
        parent_num = row.get('parent_num')
        if parent_num:
            new_children_by_parent.setdefault(parent_num, []).append(row)

    added = []
    removed = []
    changed = []

    for key, new_row in new_index.items():
        old_row = old_index.get(key)
        if not old_row:
            added.append(new_row)
            continue
        compare_row = new_row
        if old_row.get('unit_type') != new_row.get('unit_type') and new_row.get('unit_type') == 'adet':
            raw_children = new_children_by_parent.get(new_row.get('num'), [])
            raw_child = next((child for child in raw_children if child.get('unit_type') == old_row.get('unit_type')), None)
            if raw_child:
                compare_row = dict(new_row)
                compare_row.update({
                    'name': old_row.get('name'),
                    'quantity': raw_child.get('quantity'),
                    'quantity_net': raw_child.get('quantity_net'),
                    'piece_count': old_row.get('piece_count'),
                    'unit_type': raw_child.get('unit_type'),
                    'material': raw_child.get('name'),
                })
        changes = _row_changed_fields(old_row, compare_row)
        if changes:
            changed.append({'old': old_row, 'new': new_row, 'changes': changes})

    for key, old_row in old_index.items():
        if key not in new_index:
            removed.append(old_row)

    old_cost = 0.0
    try:
        tree = get_bom_tree(existing_bom_id, db)
        old_cost = sum(root.get('total_cost') or 0 for root in tree.get('roots', []))
    except Exception:
        old_cost = estimate_bom_rows_cost(old_rows)

    new_cost = estimate_bom_rows_cost(new_rows)

    return {
        'old_bom_id': existing_bom_id,
        'old_total': len(old_rows),
        'new_total': len(new_rows),
        'added': added,
        'removed': removed,
        'changed': changed,
        'raw_material_summary': _aggregate_raw_material_usage(new_rows),
        'stats': {
            'added': len(added),
            'removed': len(removed),
            'changed': len(changed),
            'old_cost': round(old_cost or 0, 2),
            'new_cost': round(new_cost or 0, 2),
            'cost_delta': round((new_cost or 0) - (old_cost or 0), 2),
        }
    }


# ---------------------------------------------------------------------------
# DB'ye Yükleme
# ---------------------------------------------------------------------------

def import_bom_to_db(parsed_rows: list[dict], bom_id: int, db, category_id: int = None, conflict_resolutions: dict = None) -> dict:
    from app.models import BomItem, BomNode, BomEdge, Product

    num_to_node_id: dict[str, int] = {}
    items_c = nodes_c = edges_c = products_c = unresolved_materials_c = 0

    # Çakışma çözümlerini hazırla (kullanıcı kararları yoksa default)
    if conflict_resolutions is None:
        conflict_resolutions = {}

    updated_products_c = 0

    for row in parsed_rows:
        code_prefix = str(row.get('code') or '')[:3]
        if code_prefix in STANDARD_PREFIXES or str(row.get('material', '')).lower() == 'standart parça':
            item_type = 'standart_parca'
        elif row['level'] == 0:
            item_type = 'mamul'
        elif row.get('is_ready_purchase') or _is_ready_purchase_text(row.get('material') or ''):
            item_type = 'hazir_parca'
        elif row.get('is_auto_hammadde'):
            item_type = 'hammadde'
        else:
            item_type = 'yarimamul'

        # --- Ağırlık / birim notu oluştur ---
        weight_val  = row.get('weight_per_unit', 0.0) or 0.0
        weight_u    = row.get('weight_unit', '') or ''
        qty_net     = row.get('quantity_net', 0.0) or 0.0
        unit_tp     = row.get('unit_type', 'adet')
        notes_parts = []
        if weight_val > 0:
            # Birim 'kg' ise miktarın kendisi ağırlık, ayrı birim ağırlığı göster
            if unit_tp == 'kg':
                notes_parts.append(f'Miktar ağırlık bazında (ölçüm birimi: {unit_tp})')
            elif unit_tp == 'metre':
                notes_parts.append(f'Birim ağırlık: {weight_val} {weight_u}/m')
            else:
                notes_parts.append(f'Birim ağırlık: {weight_val} {weight_u}')
        if qty_net > 0 and qty_net != row.get('quantity', 0):
            notes_parts.append(f'Firesiz miktar: {qty_net} {unit_tp}')
        product_notes = ' | '.join(notes_parts) or None

        # --- Product Master eşleştirme / oluşturma ---
        excel_code = row.get('code') or ''
        product = Product.query.filter_by(name=row['name']).filter(Product.code == excel_code if excel_code else True).first()
        if not product:
            potentials = Product.query.filter_by(name=row['name']).all()
            for p in potentials:
                p_code = p.code or ''
                if not excel_code or not p_code or excel_code == p_code:
                    product = p
                    break

        if not product and row.get('is_auto_hammadde'):
            product = _find_matching_raw_material(row)
        
        # Kullanıcı bu ürün için karar verdiyse kontrol et
        resolution = conflict_resolutions.get(row['name'], {})

        if not product and (not row.get('is_auto_hammadde') or _is_priceable_raw_material_name(row['name'])):
            base_code = _make_product_code(row['name'], row.get('code') or '')
            code = _unique_product_code(base_code)
            product = Product(
                code=code,
                name=row['name'],
                unit_type=row['unit_type'],
                type=item_type,
                material=(row['name'] if row.get('is_auto_hammadde') else (row.get('material') or None)),
                notes=product_notes,
                category_id=category_id if row['level'] == 0 else None
            )
            db.session.add(product)
            db.session.flush()
            products_c += 1
        elif not product and row.get('is_auto_hammadde'):
            unresolved_materials_c += 1
        elif product:
            # Mevcut ürün - kullanıcı kararlarına göre güncelle
            product_updated = False
            
            # Malzeme güncellemesi yalnızca kullanıcı seçerse yapılır.
            if resolution.get('update_material', False) and row.get('material') and row['material'] != (product.material or ''):
                product.material = row['material']
                product_updated = True
            
            # Tip güncellemesi
            if resolution.get('update_type', False):
                product.type = item_type
                product_updated = True
            
            # Birim tipi güncellemesi
            if resolution.get('update_unit', False):
                product.unit_type = row['unit_type']
                product_updated = True
            
            # Not bilgisi
            if product_notes and not product.notes:
                product.notes = product_notes
                product_updated = True
            
            if product_updated:
                updated_products_c += 1

        # --- BomItem — name bazlı unique ---
        item = BomItem.query.filter_by(name=row['name']).filter(BomItem.code == excel_code if excel_code else True).first()
        if not item:
            potentials_item = BomItem.query.filter_by(name=row['name']).all()
            for i in potentials_item:
                i_code = i.code or ''
                if not excel_code or not i_code or excel_code == i_code:
                    item = i
                    break

        if not item:
            item = BomItem(
                code=row.get('code') or None,
                name=row['name'],
                unit_type=row['unit_type'],
                type=item_type,
                product_id=product.id if product else None,
            )
            db.session.add(item)
            db.session.flush()
            items_c += 1
        else:
            if item.product_id is None and product:
                item.product_id = product.id
            # Kod varsa ve item'da eksikse güncelle (FORMAT C import)
            if row.get('code') and not item.code:
                item.code = row['code']

        q_net = row.get('quantity_net') or 0.0
        pc_c = row.get('piece_count') or 1.0
        w_val = row.get('weight_per_unit') or 0.0
        w_unt = row.get('weight_unit') or None
        node = BomNode(
            bom_id=bom_id,
            num=row['num'],
            level=row['level'],
            item_id=item.id,
            display_name=row['name'],
            quantity=Decimal(str(row['quantity'])),
            quantity_net=Decimal(str(q_net)) if q_net else None,
            piece_count=Decimal(str(pc_c)) if pc_c else 1,
            weight_per_unit=Decimal(str(w_val)) if w_val else None,
            weight_unit=w_unt,
            unit_type=row['unit_type']
        )
        db.session.add(node)
        db.session.flush()
        nodes_c += 1
        num_to_node_id[row['num']] = node.id

        parent_num = row.get('parent_num')
        parent_node_id = num_to_node_id.get(parent_num) if parent_num else None
        edge = BomEdge(bom_id=bom_id, parent_node_id=parent_node_id,
                       child_node_id=node.id,
                       quantity=Decimal(str(row['quantity'])))
        db.session.add(edge)
        edges_c += 1

    db.session.commit()
    return {
        'nodes': nodes_c, 
        'edges': edges_c, 
        'items': items_c, 
        'products': products_c,
        'updated': updated_products_c,
        'unresolved_materials': unresolved_materials_c
    }


# ---------------------------------------------------------------------------
# Ağaç Sorgulama
# ---------------------------------------------------------------------------

def get_bom_tree(bom_id: int, db) -> dict:
    from app.models import BomNode, BomEdge

    nodes = BomNode.query.filter_by(bom_id=bom_id).order_by(BomNode.num).all()
    edges = BomEdge.query.filter_by(bom_id=bom_id).all()

    if not nodes:
        return {'bom_id': bom_id, 'roots': [],
                'error': 'Bu bom_id için kayıt bulunamadı.'}

    child_to_parent: dict = {}
    child_qty: dict = {}
    for e in edges:
        child_to_parent[e.child_node_id] = e.parent_node_id
        try:    child_qty[e.child_node_id] = float(e.quantity)
        except: child_qty[e.child_node_id] = 1.0

    parent_to_children: dict = {}
    for n in nodes:
        parent_to_children.setdefault(child_to_parent.get(n.id), []).append(n.id)

    node_map = {n.id: n for n in nodes}

    def _num_key(num_str: str) -> tuple:
        """'1.2.3.' → (1, 2, 3) — sayısal sıralama için."""
        try:
            return tuple(int(p) for p in num_str.rstrip('.').split('.') if p)
        except ValueError:
            return (0,)

    def build(nid: int) -> dict:
        n = node_map[nid]
        item    = n.item
        product = item.product if item else None
        costing_product = product
        if item and item.type == 'hammadde' and (not costing_product or not (costing_product.unit_cost and costing_product.unit_cost > 0)):
            fallback_product = _find_costing_raw_material({
                'name': n.display_name or item.name,
                'unit_type': n.unit_type,
                'weight_per_unit': float(n.weight_per_unit or 0) if n.weight_per_unit else 0,
                'material': (product.material if product else None) or item.name or n.display_name or '',
                'is_auto_hammadde': True,
            }, exclude_product_id=product.id if product else None)
            if fallback_product and (not costing_product or fallback_product.unit_cost and fallback_product.unit_cost > 0):
                costing_product = fallback_product
        elif item and item.type == 'hammadde' and costing_product:
            fallback_product = _find_costing_raw_material({
                'name': n.display_name or item.name,
                'unit_type': n.unit_type,
                'weight_per_unit': float(n.weight_per_unit or 0) if n.weight_per_unit else 0,
                'material': (product.material if product else None) or item.name or n.display_name or '',
                'is_auto_hammadde': True,
            }, exclude_product_id=product.id if product else None)
            costing_code = (costing_product.code or '').upper()
            fallback_code = (fallback_product.code or '').upper() if fallback_product else ''
            if (
                fallback_product
                and fallback_product.unit_cost
                and fallback_product.unit_cost > 0
                and costing_code.startswith('3TB-')
                and not fallback_code.startswith('3TB-')
            ):
                costing_product = fallback_product
        children_ids = sorted(parent_to_children.get(nid, []),
                              key=lambda i: _num_key(node_map[i].num))
        # fireli / firesiz
        try:
            q_fireli  = float(n.quantity)     if n.quantity     else child_qty.get(n.id, 1.0)
            q_firesiz = float(n.quantity_net) if n.quantity_net else None
            w_per_unit= float(n.weight_per_unit) if n.weight_per_unit else None
        except Exception:
            q_fireli = child_qty.get(n.id, 1.0)
            q_firesiz = w_per_unit = None

        # Fire oranı (%)
        if q_firesiz and q_fireli and q_firesiz > 0:
            waste_ratio = round((q_fireli - q_firesiz) / q_firesiz * 100, 1)
        else:
            waste_ratio = None

        # item_type: eğer altında çocuk varsa 'yarimamul', yoksa item kaydından al
        has_children = bool(children_ids)
        raw_type = product.type if product and product.type else (item.type if item else 'hammadde')
        ready_purchase = _is_ready_purchase_text(
            ' '.join(
                _c(value)
                for value in [
                    n.display_name,
                    item.name if item else '',
                    product.material if product else '',
                    product.name if product else '',
                ]
            )
        )
        
        code_str = str(item.code) if (item and item.code) else (str(product.code) if product else '')
        code_prefix = code_str[:3]
        
        if code_prefix in STANDARD_PREFIXES:
            display_type = 'standart_parca'
        elif ready_purchase and n.level > 0:
            display_type = 'hazir_parca'
        elif has_children and n.level > 0:
            display_type = 'yarimamul'
        else:
            display_type = raw_type

        built_children = [build(cid) for cid in children_ids]
        is_hazir = (
            raw_type in ['hazir_parca', 'standart_parca']
            or display_type in ['hazir_parca', 'standart_parca']
            or ready_purchase
        )
        material_text = ' '.join(_c(value) for value in [
            product.material if product else '',
            product.name if product else '',
            n.display_name or '',
        ])
        if _should_cost_by_weight(material_text, n.unit_type, w_per_unit or 0):
            is_hazir = False

        if built_children and not is_hazir:
            # sum can fail if total_cost is None
            calc_total_cost = sum((c.get('total_cost') or 0.0) for c in built_children)
            
            # Eğer Lama veya hesaplaması kg üzerinden yapılan bir hammaddeyse ve weight_per_unit varsa
            # Üst kırılma da kendi qty_fireli metrajını vs uygulayabilir ancak
            # Alt kırılımların toplamı üst kırılımın 1 ADET (veya toplam) maliyetini oluşturur.
            # Maliyet firesiz miktar üzerinden hesaplanır; fire oranı sadece bilgi olarak gösterilir.
            cost_basis_qty = _cost_basis_quantity(q_fireli or 0, q_firesiz)
            calc_unit_cost = (calc_total_cost / cost_basis_qty) if cost_basis_qty and cost_basis_qty > 0 else 0.0
            calc_currency = costing_product.currency if costing_product and costing_product.currency else 'TRY'
        else:
            calc_unit_cost = costing_product.unit_cost if costing_product and costing_product.unit_cost else 0.0
            
            if is_hazir:
                p_count = float(n.piece_count) if getattr(n, 'piece_count', None) else 1.0
                calc_total_cost = calc_unit_cost * p_count
            else:
                cost_basis_qty = _cost_basis_quantity(q_fireli or 0, q_firesiz)
                if _should_cost_by_weight(material_text, n.unit_type, w_per_unit or 0):
                    cost_qty = _weight_cost_quantity(cost_basis_qty, w_per_unit or 0)
                else:
                    cost_qty = _cost_quantity_for_unit(
                        costing_product.unit_type if costing_product else n.unit_type,
                        n.unit_type,
                        cost_basis_qty,
                        float(n.piece_count) if getattr(n, 'piece_count', None) else 1.0,
                        w_per_unit or 0
                    )
                calc_total_cost = calc_unit_cost * cost_qty
            
            calc_currency = costing_product.currency if costing_product and costing_product.currency else 'TRY'

        return {
            'id': n.id, 'num': n.num, 'level': n.level,
            'name': n.display_name,
            'code': item.code if item else None,
            'quantity':   q_fireli,
            'quantity_net': q_firesiz,
            'piece_count': float(n.piece_count) if getattr(n, 'piece_count', None) else 1,
            'waste_ratio':  waste_ratio,
            'weight_per_unit': w_per_unit,
            'weight_unit': n.weight_unit or '',
            'unit': n.unit_type, 'item_id': n.item_id,
            'product_id': item.product_id if item else None,
            'material': product.material if product else (costing_product.material if costing_product else None),
            'item_type': display_type,
            'stock_qty': product.current_stock if product else 0,
            'unit_cost': calc_unit_cost,
            'currency': calc_currency,
            'total_cost': calc_total_cost,
            'children': built_children
        }

    root_ids = parent_to_children.get(None, [])
    roots = [build(r) for r in sorted(root_ids, key=lambda i: _num_key(node_map[i].num))]
    return {'bom_id': bom_id, 'roots': roots}


def get_bom_subtree(bom_id: int, node_id: int, db) -> dict:
    """
    Belirli bir düğümden başlayan alt ağacı döndür.
    
    Args:
        bom_id: BOM ID
        node_id: Başlangıç düğüm ID'si
        db: Database session
    
    Returns:
        dict: Tek bir node içeren ağaç yapısı
    """
    # Önce tüm ağacı al
    full_tree = get_bom_tree(bom_id, db)
    
    if not full_tree.get('roots'):
        return {'bom_id': bom_id, 'node': None, 'error': 'BOM bulunamadı'}
    
    # Belirli node'u bul (recursive arama)
    def find_node(nodes, target_id):
        for node in nodes:
            if node['id'] == target_id:
                return node
            if node.get('children'):
                result = find_node(node['children'], target_id)
                if result:
                    return result
        return None
    
    target_node = find_node(full_tree['roots'], node_id)
    
    if not target_node:
        return {'bom_id': bom_id, 'node': None, 'error': 'Düğüm bulunamadı'}
    
    return {
        'bom_id': bom_id,
        'node': target_node,
        'roots': [target_node]  # Excel fonksiyonu için roots formatında
    }


# ---------------------------------------------------------------------------
# Yardımcı Sorgular
# ---------------------------------------------------------------------------

def list_boms(db) -> list[dict]:
    from app.models import BomNode, BomItem, Product, Category
    from sqlalchemy import func
    result = (db.session.query(BomNode.bom_id,
               func.count(BomNode.id).label('node_count'),
               func.min(BomNode.created_at).label('created_at'))
              .group_by(BomNode.bom_id).order_by(BomNode.bom_id).all())
    boms = []
    for r in result:
        root = BomNode.query.filter_by(bom_id=r.bom_id, level=0).first()
        
        # Determine the underlying Product to display Category
        root_product = None
        category_name = None
        if root and root.item and root.item.product:
            root_product = root.item.product
            if root_product.category:
                category_name = root_product.category.name

        boms.append({
            'bom_id': r.bom_id,
            'root_name': root.display_name if root else f'BOM #{r.bom_id}',
            'product_id': root_product.id if root_product else None,
            'category_name': category_name,
            'node_count': r.node_count,
            'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else None
        })
    return boms


def next_bom_id(db) -> int:
    from app.models import BomNode
    from sqlalchemy import func
    return (db.session.query(func.max(BomNode.bom_id)).scalar() or 0) + 1
