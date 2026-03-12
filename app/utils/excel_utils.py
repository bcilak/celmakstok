"""
Excel/CSV Import/Export Utility Fonksiyonları
ÇELMAK Stok Takip Sistemi
"""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_product_template_simple():
    """
    Basitleştirilmiş ürün import şablonu (Kategori ID yok, web'de seçilecek)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"

    # Başlıklar
    headers = [
        'Ürün Kodu*',
        'Ürün Adı*',
        'Birim Tipi',
        'Mevcut Stok',
        'Minimum Stok',
        'Barkod',
        'Notlar'
    ]

    # Başlık satırını yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Örnek veri
    example_data = [
        ['HM-001', 'Çelik Levha 2mm', 'm2', '50', '10', '', 'Hammadde'],
        ['HM-002', 'Paslanmaz Boru 25mm', 'metre', '120', '20', '5901234123457', ''],
        ['BE-100', 'Somun M8', 'adet', '500', '100', '', 'Bağlantı'],
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Kolon genişliklerini ayarla
    column_widths = [15, 30, 12, 12, 12, 15, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Açıklama sayfası ekle
    ws_info = wb.create_sheet("Bilgi")
    info_text = [
        ["ÇELMAK STOK TAKİP SİSTEMİ - ÜRÜN İMPORT ŞABLONU (Basitleştirilmiş)", ""],
        ["", ""],
        ["Önemli:", ""],
        ["Kategorileri Excel'de yazmanıza gerek yok!", ""],
        ["Yükledikten sonra web arayüzünden her ürün için kategori seçeceksiniz.", ""],
        ["", ""],
        ["* ile işaretli alanlar zorunludur", ""],
        ["", ""],
        ["Alan Açıklamaları:", ""],
        ["Ürün Kodu*", "Ürününüzün benzersiz kodu (örn: HM-001, BE-045, YM-100)"],
        ["Ürün Adı*", "Ürünün tam adı"],
        ["Birim Tipi", "adet, kg, metre, litre, m2, m3 vb. (varsayılan: adet)"],
        ["Mevcut Stok", "Başlangıç stok miktarı (sayısal değer, varsayılan: 0)"],
        ["Minimum Stok", "Kritik stok seviyesi (sayısal değer, varsayılan: 0)"],
        ["Barkod", "Ürün barkodu (opsiyonel)"],
        ["Notlar", "Ek bilgiler (opsiyonel)"],
    ]

    for row_num, row_data in enumerate(info_text, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="FF0000")
            elif ":" in str(value) or "!" in str(value):
                cell.font = Font(bold=True)

    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 60

    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_product_template():
    """
    Ürün import için Excel şablonu oluştur
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürün Şablonu"

    # Başlıklar
    headers = [
        'Ürün Kodu*',
        'Ürün Adı*',
        'Kategori ID*',
        'Birim Tipi',
        'Mevcut Stok',
        'Minimum Stok',
        'Barkod',
        'Notlar'
    ]

    # Başlık satırını yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Örnek veri
    example_data = [
        ['ORN-001', 'Örnek Ürün 1', '1', 'adet', '100', '10', '', 'Örnek not'],
        ['ORN-002', 'Örnek Ürün 2', '1', 'kg', '50', '5', '1234567890', ''],
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Kolon genişliklerini ayarla
    column_widths = [15, 30, 12, 12, 12, 12, 15, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Açıklama sayfası ekle
    ws_info = wb.create_sheet("Bilgi")
    info_text = [
        ["ÇELMAK STOK TAKİP SİSTEMİ - ÜRÜN İMPORT ŞABLONU", ""],
        ["", ""],
        ["Açıklama:", ""],
        ["* ile işaretli alanlar zorunludur", ""],
        ["", ""],
        ["Alan Açıklamaları:", ""],
        ["Ürün Kodu", "Ürününüzün benzersiz kodu (örn: HM-001, OS-123)"],
        ["Ürün Adı", "Ürünün tam adı"],
        ["Kategori ID", "Kategorinin ID numarası (Kategoriler sayfasından bakabilirsiniz)"],
        ["Birim Tipi", "adet, kg, metre, litre vb."],
        ["Mevcut Stok", "Başlangıç stok miktarı (sayısal değer)"],
        ["Minimum Stok", "Kritik stok seviyesi (sayısal değer)"],
        ["Barkod", "Ürün barkodu (opsiyonel)"],
        ["Notlar", "Ek bilgiler (opsiyonel)"],
    ]

    for row_num, row_data in enumerate(info_text, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif ":" in str(value):
                cell.font = Font(bold=True)

    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 60

    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def parse_product_excel_simple(file_stream):
    """
    Basitleştirilmiş ürün Excel'ini parse et (Kategori ID yok)

    Returns:
        tuple: (success_list, error_list)
    """
    try:
        # Önce 'Ürünler' sheet'ini dene, yoksa ilk sheet'i al
        try:
            df = pd.read_excel(file_stream, sheet_name='Ürünler')
        except ValueError:
            # 'Ürünler' sheet'i yoksa, ilk sheet'i kullan
            df = pd.read_excel(file_stream, sheet_name=0)

        # Boş satırları temizle
        df = df.dropna(how='all')

        # Örnek satırları atla - SADECE tam eşleşenler
        example_codes = ['HM-001', 'HM-002', 'BE-100']
        df = df[~df['Ürün Kodu*'].astype(str).isin(example_codes)]

        success_list = []
        error_list = []

        for index, row in df.iterrows():
            try:
                # Zorunlu alanları kontrol et
                if pd.isna(row['Ürün Kodu*']) or pd.isna(row['Ürün Adı*']):
                    error_list.append({
                        'row': index + 2,
                        'error': 'Zorunlu alanlar boş olamaz (Ürün Kodu, Ürün Adı)'
                    })
                    continue

                product_data = {
                    'code': str(row['Ürün Kodu*']).strip(),
                    'name': str(row['Ürün Adı*']).strip(),
                    'unit_type': str(row.get('Birim Tipi', 'adet')).strip() if not pd.isna(row.get('Birim Tipi')) else 'adet',
                    'current_stock': float(row.get('Mevcut Stok', 0)) if not pd.isna(row.get('Mevcut Stok')) else 0,
                    'minimum_stock': float(row.get('Minimum Stok', 0)) if not pd.isna(row.get('Minimum Stok')) else 0,
                    'barcode': str(row.get('Barkod', '')).strip() if not pd.isna(row.get('Barkod')) else None,
                    'notes': str(row.get('Notlar', '')).strip() if not pd.isna(row.get('Notlar')) else None,
                    'category_id': None  # Web'de seçilecek
                }

                success_list.append(product_data)

            except Exception as e:
                error_list.append({
                    'row': index + 2,
                    'error': f'Hata: {str(e)}'
                })

        return success_list, error_list

    except Exception as e:
        return [], [{'row': 0, 'error': f'Dosya okuma hatası: {str(e)}'}]


def parse_product_excel(file_stream):
    """
    Ürün Excel dosyasını parse et

    Returns:
        tuple: (success_list, error_list)
    """
    try:
        # Önce 'Ürün Şablonu' sheet'ini dene, yoksa ilk sheet'i al
        try:
            df = pd.read_excel(file_stream, sheet_name='Ürün Şablonu')
        except ValueError:
            # 'Ürün Şablonu' sheet'i yoksa, ilk sheet'i kullan
            df = pd.read_excel(file_stream, sheet_name=0)

        # Boş satırları temizle
        df = df.dropna(how='all')

        # Örnek satırları atla
        df = df[~df['Ürün Kodu*'].astype(str).str.startswith('ORN-')]

        success_list = []
        error_list = []

        for index, row in df.iterrows():
            try:
                # Zorunlu alanları kontrol et
                if pd.isna(row['Ürün Kodu*']) or pd.isna(row['Ürün Adı*']) or pd.isna(row['Kategori ID*']):
                    error_list.append({
                        'row': index + 2,
                        'error': 'Zorunlu alanlar boş olamaz (Ürün Kodu, Ürün Adı, Kategori ID)'
                    })
                    continue

                product_data = {
                    'code': str(row['Ürün Kodu*']).strip(),
                    'name': str(row['Ürün Adı*']).strip(),
                    'category_id': int(row['Kategori ID*']),
                    'unit_type': str(row.get('Birim Tipi', 'adet')).strip() if not pd.isna(row.get('Birim Tipi')) else 'adet',
                    'current_stock': float(row.get('Mevcut Stok', 0)) if not pd.isna(row.get('Mevcut Stok')) else 0,
                    'minimum_stock': float(row.get('Minimum Stok', 0)) if not pd.isna(row.get('Minimum Stok')) else 0,
                    'barcode': str(row.get('Barkod', '')).strip() if not pd.isna(row.get('Barkod')) else None,
                    'notes': str(row.get('Notlar', '')).strip() if not pd.isna(row.get('Notlar')) else None,
                }

                success_list.append(product_data)

            except Exception as e:
                error_list.append({
                    'row': index + 2,
                    'error': f'Hata: {str(e)}'
                })

        return success_list, error_list

    except Exception as e:
        return [], [{'row': 0, 'error': f'Dosya okuma hatası: {str(e)}'}]


def export_products_to_excel(products):
    """
    Ürünleri Excel'e dışa aktar
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"

    # Başlıklar
    headers = [
        'ID',
        'Ürün Kodu',
        'Ürün Adı',
        'Kategori',
        'Birim',
        'Mevcut Stok',
        'Minimum Stok',
        'Durum',
        'Barkod',
        'Notlar',
        'Oluşturulma Tarihi'
    ]

    # Başlık satırını yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Veri satırlarını yaz
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.id)
        ws.cell(row=row_num, column=2, value=product.code)
        ws.cell(row=row_num, column=3, value=product.name)
        ws.cell(row=row_num, column=4, value=product.category.name if product.category else '')
        ws.cell(row=row_num, column=5, value=product.unit_type)
        ws.cell(row=row_num, column=6, value=product.current_stock)
        ws.cell(row=row_num, column=7, value=product.minimum_stock)

        # Stok durumu
        if product.current_stock <= 0:
            status = 'BOŞ'
            status_color = 'FF0000'
        elif product.minimum_stock > 0 and product.current_stock < product.minimum_stock:
            status = 'KRİTİK'
            status_color = 'FFA500'
        else:
            status = 'NORMAL'
            status_color = '00FF00'

        cell = ws.cell(row=row_num, column=8, value=status)
        cell.font = Font(bold=True, color=status_color)

        ws.cell(row=row_num, column=9, value=product.barcode or '')
        ws.cell(row=row_num, column=10, value=product.notes or '')
        ws.cell(row=row_num, column=11, value=product.created_at.strftime('%Y-%m-%d %H:%M') if product.created_at else '')

    # Kolon genişliklerini ayarla
    column_widths = [8, 15, 30, 20, 10, 12, 12, 10, 15, 30, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def export_stock_movements_to_excel(movements):
    """
    Stok hareketlerini Excel'e dışa aktar
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Hareketleri"

    # Başlıklar
    headers = [
        'ID',
        'Tarih',
        'Ürün Kodu',
        'Ürün Adı',
        'Kategori',
        'Hareket Tipi',
        'Miktar',
        'Birim',
        'Önceki Stok',
        'Yeni Stok',
        'Kullanıcı',
        'Açıklama'
    ]

    # Başlık satırını yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Veri satırlarını yaz
    for row_num, movement in enumerate(movements, 2):
        ws.cell(row=row_num, column=1, value=movement.id)
        ws.cell(row=row_num, column=2, value=movement.date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=3, value=movement.product.code if movement.product else '')
        ws.cell(row=row_num, column=4, value=movement.product.name if movement.product else '')
        ws.cell(row=row_num, column=5, value=movement.product.category.name if movement.product and movement.product.category else '')
        ws.cell(row=row_num, column=6, value=movement.movement_type)
        ws.cell(row=row_num, column=7, value=movement.quantity)
        ws.cell(row=row_num, column=8, value=movement.product.unit_type if movement.product else '')
        ws.cell(row=row_num, column=9, value=movement.previous_stock)
        ws.cell(row=row_num, column=10, value=movement.new_stock)
        ws.cell(row=row_num, column=11, value=movement.user.name if movement.user else '')
        ws.cell(row=row_num, column=12, value=movement.description or '')

    # Kolon genişliklerini ayarla
    column_widths = [8, 16, 15, 30, 20, 15, 10, 10, 12, 12, 15, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def parse_bom_excel(file_stream, main_product_name):
    """
    Parse the provided Urun Agaci (BOM) Excel template.
    Returns a list of components with their parent relationships.
    """
    try:
        try:
            # header=None prevents pandas from using the first row as column names,
            # which is critical because some users put top-level BOM groups (like Kasa) on row 1
            df = pd.read_excel(file_stream, sheet_name=0, header=None)
        except Exception as e:
            return [], [{'row': 0, 'error': f'Dosya okuma hatası: {str(e)}'}]
        
        success_list = []
        error_list = []
        
        current_group_level1 = None
        current_group_level2 = None
        
        for index, row in df.iterrows():
            # DO NOT SKIP index 0 anymore, as it might contain valid groups.
            
            r = [str(x).strip() if pd.notna(x) else '' for x in row.tolist()]
            if not any(bool(x) for x in r): continue
            
            col0 = r[0] if len(r) > 0 else '' # Group Level 1
            col1 = r[1] if len(r) > 1 else '' # Group Level 2 / Or Item if others empty
            col2 = r[2] if len(r) > 2 else '' 
            col3 = r[3] if len(r) > 3 else '' # Real Item Name
            col4 = r[4] if len(r) > 4 else '' # Item Material/Size
            col5 = r[5] if len(r) > 5 else '' # Qty
            col6 = r[6] if len(r) > 6 else '' # Unit
            col7 = r[7] if len(r) > 7 else '' # Main Qty
            col8 = r[8] if len(r) > 8 else '' # Main Unit
            
            # Detect group hierarchies
            if col0 and col0.lower() != 'nan':
                if current_group_level1 != col0:
                    current_group_level1 = col0
                    current_group_level2 = None
                    # Level 1 group is a child of the MAIN product
                    success_list.append({
                        'parent_name': main_product_name,
                        'name': current_group_level1,
                        'quantity': 1.0, # Excel doesn't always specify, default to 1
                        'unit_type': 'adet'
                    })
                    
            if col1 and col1.lower() != 'nan':
                # Level 2 Group detection:
                # Case 1: Pure group header (no item details in col3, col4, col5, col7)
                # Case 2: Group header + first item on the same line (col1 has text AND col3 has text)
                is_group = False
                if not col3 and not col4 and not col5 and not col7:
                    is_group = True
                elif col1 and col3:
                    is_group = True
                    
                if is_group:
                    if current_group_level2 != col1:
                        current_group_level2 = col1
                        # Level 2 group is a child of the Level 1 group (or MAIN if Level 1 is missing, though unlikely)
                        parent_of_l2 = current_group_level1 if current_group_level1 else main_product_name
                        success_list.append({
                            'parent_name': parent_of_l2,
                            'name': current_group_level2,
                            'quantity': 1.0,
                            'unit_type': 'adet'
                        })
            
            # Parent designation for the CURRENT ITEM
            parent_name = current_group_level2 if current_group_level2 else (current_group_level1 if current_group_level1 else main_product_name)
            
            # Item designation
            item_name = ''
            if col3 and col3.lower() != 'nan':
                item_name = col3
            elif col2 and col2.lower() != 'nan':
                item_name = col2
            elif col1 and col1.lower() != 'nan' and not col3:
                # Fallback if no specific item column is used, but col1 has data and it's not a pure group
                if col4 or col5 or col7: 
                    item_name = col1
            
            if not item_name: continue
            
            # QTY and UNIT
            qty_str = col7 if col7 else col5 # Prefer the "Main QTY" (Ana Adet) column
            unit_str = col8 if col8 else col6
            
            # Combine material info if exists
            if col4 and col4.lower() != 'nan':
                # Sadece eğer aynı isimde geçmiyorsa ekle
                if col4 not in item_name:
                    item_name += f' ({col4})'
                    
            qty = 1.0
            try:
                qty = float(qty_str.replace(',','.'))
            except:
                if col5:
                    try: qty = float(col5.replace(',','.'))
                    except: pass
                    
            unit = 'adet'
            if 'kilo' in unit_str.lower() or 'kg' in unit_str.lower():
                unit = 'kg'
            elif 'metre' in unit_str.lower() or 'mt' in unit_str.lower():
                unit = 'metre'
                
            success_list.append({
                'parent_name': parent_name,
                'name': item_name,
                'quantity': qty,
                'unit_type': unit
            })
                
        return success_list, error_list
    except Exception as e:
        return [], [{'row': 0, 'error': f'Beklenmeyen hata: {str(e)}'}]


def create_bom_tree_excel(tree_data: dict, bom_id: int, node_info: dict = None) -> BytesIO:
    """
    BOM ağaç yapısını Excel dosyası olarak oluşturur.
    
    Args:
        tree_data: get_bom_tree() fonksiyonundan dönen ağaç verisi veya tek bir node
        bom_id: BOM ID numarası
        node_info: Belirli bir düğüm seçilmişse {'id': node_id, 'num': num, 'name': name}
    
    Returns:
        BytesIO: Excel dosyası içeren BytesIO nesnesi
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"BOM #{bom_id}"
    
    # Başlık satırları
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    if node_info:
        title_cell.value = f'BOM #{bom_id} - {node_info.get("num", "")} {node_info.get("name", "Alt Ağaç")}'
    else:
        title_cell.value = f'BOM #{bom_id} - Ürün Ağacı'
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:N2')
    date_cell = ws['A2']
    date_cell.value = f'Oluşturma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center")
    
    # Kolon başlıkları
    headers = [
        'No',
        'Ürün Adı',
        'Ürün Kodu',
        'Malzeme',
        'Tür',
        'Miktar (Fireli)',
        'Miktar (Firesiz)',
        'Fire %',
        'Birim',
        'Uzunluk Fireli (m)',
        'Uzunluk Firesiz (m)',
        'Ağırlık Fireli (kg)',
        'Ağırlık Firesiz (kg)',
        'Stok'
    ]
    
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Kenarlıklar
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Kolon genişlikleri
    column_widths = [8, 35, 15, 20, 15, 15, 15, 10, 10, 16, 16, 16, 16, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Satır yüksekliği
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 20
    
    # Ağaç verisini düzleştir ve Excel'e yaz
    current_row = header_row + 1
    
    def flatten_and_write(node: dict, indent_level: int = 0):
        """
        Ağaç düğümünü düzleştirip Excel'e yaz (recursive)
        """
        nonlocal current_row
        
        # Hücre kenarlıkları
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # No (sıra numarası)
        num_cell = ws.cell(row=current_row, column=1)
        num_cell.value = node.get('num', '')
        num_cell.alignment = Alignment(horizontal="center")
        num_cell.border = thin_border
        num_cell.font = Font(bold=True, color="2563eb")
        
        # Ürün Adı (girintili)
        name_cell = ws.cell(row=current_row, column=2)
        indent = "  " * indent_level  # Her seviye için 2 boşluk
        name_cell.value = f"{indent}{node.get('name', '')}"
        name_cell.border = thin_border
        
        # Seviye bazlı renklendirme
        if indent_level == 0:
            # Ana ürün - Koyu sarı
            name_cell.fill = PatternFill(start_color="fef08a", end_color="fef08a", fill_type="solid")
            name_cell.font = Font(bold=True)
        elif indent_level == 1:
            # 1. seviye - Açık mavi
            name_cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        elif indent_level == 2:
            # 2. seviye - Açık yeşil
            name_cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        
        # Ürün Kodu
        code_cell = ws.cell(row=current_row, column=3)
        code_cell.value = node.get('code', '')
        code_cell.border = thin_border
        
        # Malzeme
        material_cell = ws.cell(row=current_row, column=4)
        material_cell.value = node.get('material', '')
        material_cell.border = thin_border
        
        # Tür
        type_cell = ws.cell(row=current_row, column=5)
        item_type = node.get('item_type', '')
        
        # Türkçe karşılıklar
        type_map = {
            'yarimamul': 'Yarı Mamul',
            'hammadde': 'Hammadde',
            'mamul': 'Mamul',
            'standart_parca': 'Standart Parça'
        }
        type_cell.value = type_map.get(item_type, item_type)
        type_cell.border = thin_border
        
        # Tür bazlı renklendirme
        if item_type == 'yarimamul':
            type_cell.fill = PatternFill(start_color="fff8e1", end_color="fff8e1", fill_type="solid")
            type_cell.font = Font(color="f57f17")
        elif item_type == 'hammadde':
            type_cell.fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
            type_cell.font = Font(color="880e4f")
        elif item_type == 'standart_parca':
            type_cell.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
            type_cell.font = Font(color="0d47a1")
        
        # Miktar (Fireli)
        qty_fireli_cell = ws.cell(row=current_row, column=6)
        qty_fireli = node.get('quantity', '')
        if qty_fireli and qty_fireli != '':
            try:
                qty_fireli_cell.value = float(qty_fireli)
                qty_fireli_cell.number_format = '0.####'
            except:
                qty_fireli_cell.value = qty_fireli
        qty_fireli_cell.alignment = Alignment(horizontal="right")
        qty_fireli_cell.border = thin_border
        qty_fireli_cell.fill = PatternFill(start_color="fff1f2", end_color="fff1f2", fill_type="solid")
        
        # Miktar (Firesiz)
        qty_firesiz_cell = ws.cell(row=current_row, column=7)
        qty_firesiz = node.get('quantity_net', '')
        if qty_firesiz and qty_firesiz != '':
            try:
                qty_firesiz_cell.value = float(qty_firesiz)
                qty_firesiz_cell.number_format = '0.####'
            except:
                qty_firesiz_cell.value = qty_firesiz
        qty_firesiz_cell.alignment = Alignment(horizontal="right")
        qty_firesiz_cell.border = thin_border
        qty_firesiz_cell.fill = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
        
        # Fire %
        waste_cell = ws.cell(row=current_row, column=8)
        waste_ratio = node.get('waste_ratio', '')
        if waste_ratio and waste_ratio != '':
            try:
                waste_cell.value = float(waste_ratio)
                waste_cell.number_format = '0.#"%"'
            except:
                waste_cell.value = waste_ratio
        waste_cell.alignment = Alignment(horizontal="right")
        waste_cell.border = thin_border
        
        # Birim
        unit_cell = ws.cell(row=current_row, column=9)
        unit_cell.value = node.get('unit', '')
        unit_cell.alignment = Alignment(horizontal="center")
        unit_cell.border = thin_border
        
        # Uzunluk ve Ağırlık hesaplamaları
        unit_type = node.get('unit', '')
        qty_fireli_val = node.get('quantity', 0)
        qty_firesiz_val = node.get('quantity_net', 0)
        weight_per_unit = node.get('weight_per_unit', 0)
        
        # Uzunluk (Fireli) - Sadece metre biriminde
        length_fireli_cell = ws.cell(row=current_row, column=10)
        if unit_type == 'metre' and qty_fireli_val:
            try:
                length_fireli_cell.value = float(qty_fireli_val)
                length_fireli_cell.number_format = '0.####'
            except:
                pass
        length_fireli_cell.alignment = Alignment(horizontal="right")
        length_fireli_cell.border = thin_border
        length_fireli_cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        
        # Uzunluk (Firesiz) - Sadece metre biriminde
        length_firesiz_cell = ws.cell(row=current_row, column=11)
        if unit_type == 'metre' and qty_firesiz_val:
            try:
                length_firesiz_cell.value = float(qty_firesiz_val)
                length_firesiz_cell.number_format = '0.####'
            except:
                pass
        length_firesiz_cell.alignment = Alignment(horizontal="right")
        length_firesiz_cell.border = thin_border
        length_firesiz_cell.fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
        
        # Ağırlık (Fireli) - kg biriminde veya hesaplanan
        weight_fireli_cell = ws.cell(row=current_row, column=12)
        if unit_type == 'kg' and qty_fireli_val:
            try:
                weight_fireli_cell.value = float(qty_fireli_val)
                weight_fireli_cell.number_format = '0.####'
            except:
                pass
        elif weight_per_unit and weight_per_unit > 0 and qty_fireli_val:
            try:
                total_weight = float(weight_per_unit) * float(qty_fireli_val)
                weight_fireli_cell.value = total_weight
                weight_fireli_cell.number_format = '0.####'
            except:
                pass
        weight_fireli_cell.alignment = Alignment(horizontal="right")
        weight_fireli_cell.border = thin_border
        weight_fireli_cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        
        # Ağırlık (Firesiz) - kg biriminde veya hesaplanan
        weight_firesiz_cell = ws.cell(row=current_row, column=13)
        if unit_type == 'kg' and qty_firesiz_val:
            try:
                weight_firesiz_cell.value = float(qty_firesiz_val)
                weight_firesiz_cell.number_format = '0.####'
            except:
                pass
        elif weight_per_unit and weight_per_unit > 0 and qty_firesiz_val:
            try:
                total_weight = float(weight_per_unit) * float(qty_firesiz_val)
                weight_firesiz_cell.value = total_weight
                weight_firesiz_cell.number_format = '0.####'
            except:
                pass
        weight_firesiz_cell.alignment = Alignment(horizontal="right")
        weight_firesiz_cell.border = thin_border
        weight_firesiz_cell.fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
        
        # Stok
        stock_cell = ws.cell(row=current_row, column=14)
        stock_qty = node.get('stock_qty', 0)
        try:
            stock_cell.value = float(stock_qty)
            stock_cell.number_format = '0.####'
        except:
            stock_cell.value = stock_qty
        stock_cell.alignment = Alignment(horizontal="right")
        stock_cell.border = thin_border
        
        current_row += 1
        
        # Alt düğümleri işle
        for child in node.get('children', []):
            flatten_and_write(child, indent_level + 1)
    
    # Tüm kök düğümleri işle
    roots = tree_data.get('roots', [])
    for root in roots:
        flatten_and_write(root, 0)
    
    # Alt bilgi
    info_row = current_row + 2
    ws.merge_cells(f'A{info_row}:N{info_row}')
    info_cell = ws[f'A{info_row}']
    info_cell.value = '© ÇELMAK Stok Takip Sistemi - BOM Ağaç Raporu'
    info_cell.font = Font(size=9, italic=True, color="64748b")
    info_cell.alignment = Alignment(horizontal="center")
    
    # Özet Sayfa Ekle
    ws_summary = wb.create_sheet("Özet")
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 50
    
    summary_data = [
        ["BOM ÖZET BİLGİLERİ", ""],
        ["", ""],
        ["BOM ID:", bom_id],
        ["Oluşturma Tarihi:", datetime.now().strftime("%d.%m.%Y %H:%M:%S")],
        ["Toplam Kök Sayısı:", len(roots)],
        ["", ""],
        ["AÇIKLAMALAR:", ""],
        ["Fireli Miktar:", "Fire dahil gerekli miktar"],
        ["Firesiz Miktar:", "Net (fire hariç) miktar"],
        ["Fire %:", "Fire oranı (Fireli - Firesiz) / Firesiz * 100"],
        ["Uzunluk:", "Sadece birim 'metre' olan ürünlerde görünür"],
        ["Ağırlık:", "Birim 'kg' ise miktar, değilse birim ağırlık × miktar"],
        ["Tür:", "Ürün tipi (Hammadde, Yarı Mamul, Mamul, Standart Parça)"],
        ["", ""],
        ["RENK KODLARI:", ""],
        ["Ana Ürün:", "Koyu sarı arka plan"],
        ["1. Seviye:", "Açık mavi arka plan"],
        ["2. Seviye:", "Açık yeşil arka plan"],
        ["Uzunluk Kolonları:", "Sarı tonlarda"],
        ["Ağırlık Kolonları:", "Mavi tonlarda"],
    ]
    
    for row_num, (key, value) in enumerate(summary_data, 1):
        key_cell = ws_summary.cell(row=row_num, column=1)
        key_cell.value = key
        value_cell = ws_summary.cell(row=row_num, column=2)
        value_cell.value = value
        
        if row_num == 1:
            ws_summary.merge_cells(f'A{row_num}:B{row_num}')
            key_cell.font = Font(bold=True, size=14, color="FFFFFF")
            key_cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            key_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_summary.row_dimensions[row_num].height = 25
        elif ":" in key:
            key_cell.font = Font(bold=True)
    
    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

