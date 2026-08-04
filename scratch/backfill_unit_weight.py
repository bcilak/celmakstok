"""
Product.unit_weight (1 adet = kaç kg/metre) alanını Excel sayım listesindeki
KG / METRE sütunlarından doldurur.

Kullanım (proje kök dizininde):
    python scratch/backfill_unit_weight.py "C:/.../01.06.2026 SAYIM LISTESI.xlsx"          # DRY-RUN (sadece rapor)
    python scratch/backfill_unit_weight.py "C:/.../01.06.2026 SAYIM LISTESI.xlsx" --apply   # veritabanina yaz

Mantik:
- Excel 'Hammaddeler' sayfasi: A=Malzeme adi, G=KG (1 adet kac kg), H=METRE (1 adet kac metre).
- Kartin unit_type'i kg/gr/ton ise -> unit_weight = G (kg/adet).
- unit_type metre/mt ise -> unit_weight = H (metre/adet).
- unit_type adet ise -> cevrim yok, atlanir.
- Eslestirme: urun adi normalize edilerek (kucuk harf, bosluk sadelestirme) yapilir.
"""
import sys
import re
import openpyxl

sys.path.insert(0, '.')
from app import create_app, db
from app.models import Product


def norm(s):
    """Isim normalizasyonu: kucuk harf + fazla bosluklari sadelestir."""
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def load_excel_factors(path):
    """Excel 'Hammaddeler' sayfasindan {normalize_ad: (kg_per_adet, metre_per_adet)} dondurur."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Hammaddeler']
    factors = {}
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 1).value           # A: Malzeme adi
        g = ws.cell(r, 7).value              # G: KG (kg/adet)
        h = ws.cell(r, 8).value              # H: METRE (metre/adet)
        if not name:
            continue
        kg = float(g) if isinstance(g, (int, float)) else 0.0
        mt = float(h) if isinstance(h, (int, float)) else 0.0
        factors[norm(name)] = (kg, mt)
    return factors


def main():
    if len(sys.argv) < 2:
        print("Kullanim: python scratch/backfill_unit_weight.py <excel_yolu> [--apply]")
        sys.exit(1)
    excel_path = sys.argv[1]
    apply = '--apply' in sys.argv

    factors = load_excel_factors(excel_path)
    print(f"Excel'den {len(factors)} malzeme katsayisi okundu.\n")

    app = create_app()
    with app.app_context():
        products = Product.query.all()
        matched = updated = skipped_adet = unmatched = 0
        report = []
        for p in products:
            unit = (p.unit_type or '').lower()
            if unit in ('adet', ''):
                skipped_adet += 1
                continue
            f = factors.get(norm(p.name))
            if not f:
                unmatched += 1
                continue
            kg, mt = f
            new_val = kg if unit in ('kg', 'gr', 'ton') else (mt if unit in ('metre', 'mt') else 0)
            if not new_val:
                unmatched += 1
                continue
            matched += 1
            if p.unit_weight != new_val:
                report.append(f"  {p.code} | {p.name[:40]:40} | {unit} | {p.unit_weight} -> {new_val}")
                if apply:
                    p.unit_weight = new_val
                updated += 1
        if apply:
            db.session.commit()

        print("\n".join(report[:60]))
        if len(report) > 60:
            print(f"  ... (+{len(report) - 60} satir daha)")
        print("\n--- OZET ---")
        print(f"Eslesen (kg/metre karti)  : {matched}")
        print(f"Guncellenecek/guncellenen : {updated}")
        print(f"Adet karti (atlandi)      : {skipped_adet}")
        print(f"Eslesmeyen                : {unmatched}")
        print(f"\nMod: {'YAZILDI (--apply)' if apply else 'DRY-RUN (sadece rapor). Yazmak icin --apply ekleyin.'}")


if __name__ == '__main__':
    main()
